#!/usr/bin/env python3
"""Import a BMG Research poll directly from an XLSX URL."""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import select

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from db import Database
from models import Poll, PollRow, Pollster
from polls.import_types import PollImportResult

DEFAULT_XLSX_URL = "https://bmgresearch.com/wp-content/uploads/2026/02/january-2026-omnibus-tables-for-the-i.xlsx"
DEFAULT_MAP_NAME = "UK Constituencies post 2022"
DEFAULT_POLLSTER_IDENTIFIER = "bmg_research"
NATIONAL_KEY = "__national__"

SOURCE_REGION_TO_INTERNAL = {
    "East Midlands": "East Midlands",
    "East of England": "East of England",
    "London": "London",
    "North East": "North East England",
    "North West": "North West England",
    "South East": "South East England",
    "South West": "South West England",
    "West Midlands": "West Midlands",
    "Yorkshire and The Humber": "Yorkshire and The Humber",
    "Scotland": "Scotland",
    "Wales": "Wales",
}

PARTY_NAME_MAP = {
    "Conservative": "Conservative",
    "Conservatives": "Conservative",
    "Labour": "Labour",
    "Liberal Democrats": "Liberal Democrats",
    "Liberal Democrat": "Liberal Democrats",
    "Reform UK": "Reform UK",
    "Green": "Green",
    "Green Party": "Green",
    "Scottish National Party": "Scottish National Party",
    "Scottish National Party (SNP)": "Scottish National Party",
    "SNP": "Scottish National Party",
    "Plaid Cymru": "Plaid Cymru",
    "Another Party": "Other",
    "Another Party / An independent candidate": "Other",
    "Another Party / An Independent Candidate": "Other",
    "Another Party/An independent candidate": "Other",
    "Other": "Other",
}


class ParsedPoll(BaseModel):
    """Parsed poll data extracted from source."""

    sample_size: int = Field(gt=0)
    fieldwork_start: date
    fieldwork_end: date
    party_region_percentages: dict[str, dict[str, float]]


class PlannedPollRow(BaseModel):
    """A single poll row planned for DB insertion."""

    party_id: int
    party_name: str
    region_id: int | None
    region_name: str
    percentage: float = Field(ge=0, le=100)


class ImportPlan(BaseModel):
    """Full import plan: pollster, poll metadata, and rows."""

    pollster_identifier: str
    pollster_name: str
    pollster_id: int | None
    pollster_exists: bool
    regions_mapping: str
    map_id: int
    map_name: str
    source_url: str
    parsed: ParsedPoll
    poll_id: int | None
    poll_exists: bool
    rows: list[PlannedPollRow]


def _month_number(month_text: str) -> int | None:
    """Return the integer month number for a month name string, or None if unrecognised.

    Args:
        month_text: A month name or abbreviation, e.g. ``"January"``, ``"jan"``, ``"Sept."``.
            Case-insensitive; trailing periods are stripped.

    Returns:
        Integer month number in the range 1–12, or ``None`` if the text is not recognised.
    """
    month_map = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }
    return month_map.get(month_text.strip().lower().rstrip("."))


def _infer_year(xlsx_url: str, fallback: int | None = None) -> int:
    """Infer a four-digit year from an XLSX URL, with an optional fallback.

    Looks for a ``/YYYY/`` path segment first, then any ``20xx`` substring in
    the URL.  If neither is found and ``fallback`` is provided it is returned;
    otherwise ``ValueError`` is raised.

    Args:
        xlsx_url: The URL of the XLSX file, e.g.
            ``"https://bmgresearch.com/wp-content/uploads/2026/02/tables.xlsx"``.
        fallback: Optional year to return when the URL contains no recognisable
            year.  Must be a four-digit integer if supplied.

    Returns:
        The inferred or fallback year as an integer.

    Raises:
        ValueError: If no year can be inferred from ``xlsx_url`` and no
            ``fallback`` is provided.
    """
    match = re.search(r"/(20\d{2})/", xlsx_url)
    if match:
        return int(match.group(1))
    matches = re.findall(r"(20\d{2})", xlsx_url)
    if matches:
        return int(matches[0])
    if fallback is not None:
        return fallback
    raise ValueError("Could not infer year from URL")


def _parse_fieldwork(fieldwork_text: str, default_year: int | None = None) -> tuple[date, date]:
    """Parse a human-readable fieldwork date string into a (start, end) date pair.

    Handles multiple formats found in BMG Research spreadsheets, including:

    * Same-month ampersand range: ``"3 & 4 January 2026"``
    * Same-month hyphen range: ``"3-4 January 2026"``
    * Cross-month range: ``"31 January - 2 February 2026"``
    * Range without explicit year (requires ``default_year``): ``"3-4 January"``
    * Single day with year: ``"3 January 2026"``
    * Single day without year (requires ``default_year``): ``"3 January"``

    Ordinal suffixes (``st``, ``nd``, ``rd``, ``th``) are stripped before
    parsing.  Em-dashes and en-dashes are normalised to hyphens.

    Args:
        fieldwork_text: The raw fieldwork date string as read from the
            spreadsheet, e.g. ``"Fieldwork dates: 3rd-4th January 2026"``.
        default_year: Year to use when the date string contains no explicit
            year.  Optional; required for the no-year patterns.

    Returns:
        A ``(fieldwork_start, fieldwork_end)`` tuple of ``datetime.date``
        objects.  For single-day fieldwork both elements are the same date.

    Raises:
        ValueError: If the month name cannot be parsed, or if no pattern
            matches the input string.
    """
    normalized = re.sub(r"\s+", " ", fieldwork_text.strip().replace("–", "-").replace("—", "-"))
    normalized = re.sub(r"(\d)(st|nd|rd|th)", r"\1", normalized, flags=re.IGNORECASE)

    pattern_same_month_amp = re.compile(r"(\d{1,2})\s*&\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})")
    match = pattern_same_month_amp.search(normalized)
    if match:
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        year = int(match.group(4))
        if month is None:
            raise ValueError(f"Could not parse fieldwork month: {fieldwork_text!r}")
        return date(year, month, day_start), date(year, month, day_end)

    pattern_same_month = re.compile(r"(\d{1,2})\s*-\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})")
    match = pattern_same_month.search(normalized)
    if match:
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        year = int(match.group(4))
        if month is None:
            raise ValueError(f"Could not parse fieldwork month: {fieldwork_text!r}")
        return date(year, month, day_start), date(year, month, day_end)

    pattern_cross_month = re.compile(
        r"(\d{1,2})\s+([A-Za-z]+)\s*-\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})"
    )
    match = pattern_cross_month.search(normalized)
    if match:
        day_start = int(match.group(1))
        month_start = _month_number(match.group(2))
        day_end = int(match.group(3))
        month_end = _month_number(match.group(4))
        year_end = int(match.group(5))
        if month_start is None or month_end is None:
            raise ValueError(f"Could not parse fieldwork months: {fieldwork_text!r}")
        year_start = year_end - 1 if month_start > month_end else year_end
        return date(year_start, month_start, day_start), date(year_end, month_end, day_end)

    pattern_no_year = re.compile(r"(\d{1,2})\s*-\s*(\d{1,2})\s+([A-Za-z]+)")
    match = pattern_no_year.search(normalized)
    if match and default_year is not None:
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        if month is None:
            raise ValueError(f"Could not parse fieldwork month: {fieldwork_text!r}")
        return date(default_year, month, day_start), date(default_year, month, day_end)

    pattern_single_day = re.compile(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})")
    match = pattern_single_day.search(normalized)
    if match:
        day = int(match.group(1))
        month = _month_number(match.group(2))
        year = int(match.group(3))
        if month is None:
            raise ValueError(f"Could not parse fieldwork month: {fieldwork_text!r}")
        parsed_day = date(year, month, day)
        return parsed_day, parsed_day

    pattern_single_day_no_year = re.compile(r"(\d{1,2})\s+([A-Za-z]+)$")
    match = pattern_single_day_no_year.search(normalized)
    if match and default_year is not None:
        day = int(match.group(1))
        month = _month_number(match.group(2))
        if month is None:
            raise ValueError(f"Could not parse fieldwork month: {fieldwork_text!r}")
        parsed_day = date(default_year, month, day)
        return parsed_day, parsed_day

    raise ValueError(f"Could not parse fieldwork string: {fieldwork_text!r}")


def extract_workbook(xlsx_url: str) -> Any:
    """Fetch an XLSX file from a URL and return an openpyxl ``Workbook``.

    If the primary URL uses the ``bmgresearch.co.uk`` domain, alternative
    ``bmgresearch.com`` variants are also tried in order before failing.
    Only responses whose payload begins with the ZIP magic bytes ``PK`` are
    accepted as valid XLSX content.

    Args:
        xlsx_url: URL of the XLSX file to download.  May use either the
            ``bmgresearch.co.uk`` or ``bmgresearch.com`` domain.

    Returns:
        An openpyxl ``Workbook`` instance loaded from the downloaded payload,
        with ``data_only=True`` (formula results, not formula strings).

    Raises:
        ValueError: If none of the candidate URLs returns a valid XLSX payload,
            with a combined error message listing each failure reason.
    """
    candidate_urls = [xlsx_url]
    if "bmgresearch.co.uk" in xlsx_url:
        candidate_urls.append(xlsx_url.replace("bmgresearch.co.uk", "bmgresearch.com"))
        candidate_urls.append(xlsx_url.replace("www.bmgresearch.co.uk", "bmgresearch.com"))

    payload = None
    errors: list[str] = []
    for candidate in dict.fromkeys(candidate_urls):
        req = Request(candidate, headers={"User-Agent": "Mozilla/5.0 (compatible; poll-importer/1.0)"})
        try:
            with urlopen(req, timeout=45) as response:
                data = response.read()
            if data.startswith(b"PK"):
                payload = data
                break
            errors.append(f"non-xlsx payload at {candidate}")
        except Exception as exc:
            errors.append(f"{candidate}: {exc}")

    if payload is None:
        raise ValueError("Could not fetch XLSX payload: " + " | ".join(errors))

    return load_workbook(filename=BytesIO(payload), data_only=True)


def _cell_text(value: object) -> str:
    """Return the string representation of a cell value, stripped of whitespace.

    Args:
        value: The raw cell value from openpyxl, which may be ``None``, a
            number, a string, or another type.

    Returns:
        Stripped string, or an empty string if ``value`` is ``None``.
    """
    if value is None:
        return ""
    return str(value).strip()


def _to_percentage(value: object) -> float:
    """Convert a cell value to a rounded whole-number percentage.

    Values in the range ``[0.0, 1.0]`` are assumed to be proportions and are
    multiplied by 100 before rounding.  Values outside that range are taken
    as already-percentage figures and are rounded directly.

    Args:
        value: The raw cell value to convert.  Must be numeric (``int`` or
            ``float``) and non-``None``.

    Returns:
        A float representing the percentage, rounded to the nearest integer
        (e.g. ``0.42`` → ``42.0``, ``42.4`` → ``42.0``).

    Raises:
        ValueError: If ``value`` is ``None``.
    """
    if value is None:
        raise ValueError("Encountered empty percentage cell")
    number = float(value)  # type: ignore[arg-type]
    pct = number * 100.0 if 0.0 <= number <= 1.0 else number
    return float(int(round(pct)))


def _to_percentage_or_zero(value: object) -> float:
    """Convert a cell value to a percentage, returning ``0.0`` for blank/dash cells.

    Treats ``None``, empty strings, and common dash characters (``-``, ``–``,
    ``—``) as zero.  All other values are forwarded to :func:`_to_percentage`.

    Args:
        value: The raw cell value from openpyxl.

    Returns:
        A float percentage (rounded to nearest integer), or ``0.0`` if the
        cell is blank or contains a placeholder dash.
    """
    if value is None:
        return 0.0
    if isinstance(value, str):
        cleaned = value.strip()
        if cleaned in {"", "-", "–", "—"}:
            return 0.0
    return _to_percentage(value)


def _normalize_header(value: str) -> str:
    """Collapse runs of whitespace in a header string and strip leading/trailing space.

    Args:
        value: Raw header text read from a spreadsheet cell.

    Returns:
        The header with all internal whitespace runs replaced by a single
        space and surrounding whitespace removed.
    """
    return re.sub(r"\s+", " ", value).strip()


def _find_methodology_sheet(workbook: Any) -> Any:
    """Return the methodology sheet from a workbook, falling back to the first sheet.

    Searches sheet names (case-insensitive) for one containing ``"method"``.
    If no such sheet exists, the first sheet in the workbook is returned.

    Args:
        workbook: An openpyxl ``Workbook`` instance.

    Returns:
        The openpyxl ``Worksheet`` that contains methodology metadata.
    """
    for name in workbook.sheetnames:
        if "method" in name.lower():
            return workbook[name]
    return workbook[workbook.sheetnames[0]]


def _parse_fieldwork_and_sample(workbook: Any, default_year: int) -> tuple[date, date, int]:
    """Extract fieldwork dates and sample size from the methodology sheet.

    Scans the first 80 rows and 5 columns of the methodology sheet for cells
    whose text contains ``"fieldwork date"`` (the date range) or starts with
    ``"sample:"`` (the sample size).  The sample size is extracted by stripping
    all non-digit characters from the sample line.

    Args:
        workbook: An openpyxl ``Workbook`` instance containing a methodology
            sheet (as identified by :func:`_find_methodology_sheet`).
        default_year: Fallback year passed to :func:`_parse_fieldwork` when
            the fieldwork string contains no explicit year.

    Returns:
        A ``(fieldwork_start, fieldwork_end, sample_size)`` tuple where the
        dates are ``datetime.date`` objects and ``sample_size`` is a positive
        integer.

    Raises:
        ValueError: If the fieldwork date line, sample line, or sample digits
            cannot be found in the methodology sheet.
    """
    sheet = _find_methodology_sheet(workbook)

    fieldwork_raw = None
    sample_raw = None

    for row in range(1, 80):
        for col in range(1, 6):
            text = _cell_text(sheet.cell(row, col).value)
            if not text:
                continue
            lowered = text.lower()
            if "fieldwork date" in lowered or "fieldwork dates" in lowered:
                match = re.search(r"fieldwork dates?\s*:\s*(.+)$", text, flags=re.IGNORECASE)
                fieldwork_raw = match.group(1).strip() if match else text
            if lowered.startswith("sample:"):
                sample_raw = text

    if not fieldwork_raw:
        raise ValueError("Fieldwork dates not found in methodology sheet")

    fieldwork_start, fieldwork_end = _parse_fieldwork(fieldwork_raw, default_year=default_year)

    if not sample_raw:
        raise ValueError("Sample line not found in methodology sheet")
    digits = re.sub(r"[^0-9]", "", sample_raw)
    if not digits:
        raise ValueError("Could not parse sample size from methodology sheet")

    return fieldwork_start, fieldwork_end, int(digits)


def _find_tables_sheet(workbook: Any) -> Any:
    """Return the tables sheet from a workbook.

    Matches the first sheet whose name equals ``"tables"`` (case-insensitive)
    or contains the word ``"table"``.

    Args:
        workbook: An openpyxl ``Workbook`` instance.

    Returns:
        The openpyxl ``Worksheet`` containing the cross-tabulation data.

    Raises:
        ValueError: If no sheet with a name matching the above criteria exists.
    """
    for name in workbook.sheetnames:
        if name.lower() == "tables" or "table" in name.lower():
            return workbook[name]
    raise ValueError("Could not find tables sheet")


def _find_vi_table_starts(sheet: Any) -> list[int]:
    """Return the row numbers where ``WouldVoteTodayRevised`` tables begin.

    Scans column B of the sheet for cells whose text (normalised to lower-case,
    whitespace stripped) contains ``"wouldvotetodayrevised"``.  The first match
    is treated as the national table; the last match is treated as the regional
    cross-tabulation table.

    Args:
        sheet: An openpyxl ``Worksheet`` — expected to be the tables sheet.

    Returns:
        A list of 1-based row numbers where ``WouldVoteTodayRevised`` table
        headers appear.  Typically contains two entries (national and regional).

    Raises:
        ValueError: If no matching header row is found in the sheet.
    """
    starts: list[int] = []
    for row in range(1, sheet.max_row + 1):
        value = _cell_text(sheet.cell(row, 2).value)
        if "wouldvotetodayrevised" in value.lower():
            starts.append(row)
    if not starts:
        raise ValueError("Could not find WouldVoteTodayRevised table")
    return starts


def _parse_party_region_percentages(workbook: Any) -> dict[str, dict[str, float]]:
    """Extract national and regional voting-intention percentages from the tables sheet.

    Locates the national and regional ``WouldVoteTodayRevised`` cross-tabulation
    tables, reads party rows from each, and merges them into a combined
    structure keyed first by canonical party name, then by region name (or
    :data:`NATIONAL_KEY` for the national figure).

    Party labels are normalised via :data:`PARTY_NAME_MAP`.  Regional column
    headers are mapped to internal region names via
    :data:`SOURCE_REGION_TO_INTERNAL`.  Optional parties (SNP, Plaid Cymru,
    Other) default to ``0.0`` if absent from the workbook.

    Args:
        workbook: An openpyxl ``Workbook`` instance.

    Returns:
        A dict mapping canonical party name → dict mapping region name (or
        :data:`NATIONAL_KEY`) → percentage as a rounded float.  Example::

            {
                "Labour": {
                    "__national__": 38.0,
                    "London": 52.0,
                    "Scotland": 27.0,
                    ...
                },
                ...
            }

    Raises:
        ValueError: If the tables sheet cannot be found, if no
            ``WouldVoteTodayRevised`` header is located, if no regional columns
            are found, or if any required party rows are missing from the
            national table.
    """
    sheet = _find_tables_sheet(workbook)
    starts = _find_vi_table_starts(sheet)
    national_start = starts[0]
    regional_start = starts[-1]

    national_values: dict[str, float] = {}
    for row in range(national_start + 1, min(sheet.max_row, national_start + 180)):
        label = _cell_text(sheet.cell(row, 2).value)
        if label.lower().startswith("table ") and row > national_start + 4:
            break
        canonical = PARTY_NAME_MAP.get(label)
        if canonical is None:
            continue

        next_total = sheet.cell(row + 1, 3).value
        current_total = sheet.cell(row, 3).value

        if isinstance(next_total, (int, float)) and 0.0 <= float(next_total) <= 1.2:
            percentage = _to_percentage(next_total)
        elif isinstance(current_total, (int, float)) and 0.0 <= float(current_total) <= 1.2:
            percentage = _to_percentage(current_total)
        else:
            continue

        national_values[canonical] = percentage

    region_header_row = regional_start + 3
    region_columns: dict[str, int] = {}
    for col in range(34, 90):
        header = _normalize_header(_cell_text(sheet.cell(region_header_row, col).value))
        if not header:
            continue
        internal = SOURCE_REGION_TO_INTERNAL.get(header)
        if internal is not None:
            region_columns[internal] = col

    if not region_columns:
        raise ValueError("Could not locate BMG regional columns in tables sheet")

    regional_values_by_party: dict[str, dict[str, float]] = {}
    for row in range(regional_start + 1, min(sheet.max_row, regional_start + 220)):
        label = _cell_text(sheet.cell(row, 2).value)
        if label.lower().startswith("table ") and row > regional_start + 4:
            break
        canonical = PARTY_NAME_MAP.get(label)
        if canonical is None:
            continue

        region_percentages: dict[str, float] = {}
        for region_name, col in region_columns.items():
            pct_cell = sheet.cell(row + 1, col).value
            region_percentages[region_name] = _to_percentage_or_zero(pct_cell)

        regional_values_by_party[canonical] = region_percentages

    parsed: dict[str, dict[str, float]] = {}
    all_region_names = set(region_columns.keys())
    for canonical, national_pct in national_values.items():
        merged = {NATIONAL_KEY: national_pct}
        source_regions = regional_values_by_party.get(canonical, {})
        for region_name in all_region_names:
            merged[region_name] = source_regions.get(region_name, 0.0)
        parsed[canonical] = merged

    for optional_party in ["Scottish National Party", "Plaid Cymru", "Other"]:
        if optional_party not in parsed:
            parsed[optional_party] = {NATIONAL_KEY: 0.0}
        for region_name in all_region_names:
            parsed[optional_party].setdefault(region_name, 0.0)

    required = {
        "Conservative",
        "Labour",
        "Liberal Democrats",
        "Reform UK",
        "Green",
        "Scottish National Party",
        "Plaid Cymru",
        "Other",
    }
    missing = sorted(required - set(national_values.keys()))
    if missing:
        raise ValueError(f"Missing expected party rows in workbook: {missing}")

    return parsed


def parse_poll(workbook: Any, *, source_url: str, year_hint: int | None = None) -> ParsedPoll:
    """Parse all relevant poll data from a BMG Research workbook.

    Infers the poll year from ``source_url`` (falling back to ``year_hint``),
    extracts fieldwork dates and sample size from the methodology sheet, and
    extracts national and regional voting-intention percentages from the tables
    sheet.

    Args:
        workbook: An openpyxl ``Workbook`` loaded from the BMG Research XLSX
            file, typically obtained via :func:`extract_workbook`.
        source_url: The URL from which the workbook was fetched.  Used to
            infer the poll year via :func:`_infer_year`.
        year_hint: Optional fallback year (four-digit integer) used when
            ``source_url`` contains no recognisable year and when fieldwork
            date strings omit the year.

    Returns:
        A :class:`ParsedPoll` instance containing sample size, fieldwork date
        range, and party-by-region percentage data.

    Raises:
        ValueError: If the year cannot be inferred, or if fieldwork/sample/
            party data cannot be extracted from the workbook.
    """
    inferred_year = _infer_year(source_url, fallback=year_hint)
    fieldwork_start, fieldwork_end, sample_size = _parse_fieldwork_and_sample(workbook, inferred_year)
    party_region_percentages = _parse_party_region_percentages(workbook)
    return ParsedPoll(
        sample_size=sample_size,
        fieldwork_start=fieldwork_start,
        fieldwork_end=fieldwork_end,
        party_region_percentages=party_region_percentages,
    )


def _find_existing_poll(db: Database, pollster_id: int, map_id: int, parsed: ParsedPoll) -> Poll | None:
    """Look up an existing Poll row matching the given pollster, map, and fieldwork metadata.

    Matches on ``pollster_id``, ``map_id``, ``fieldwork_start``,
    ``fieldwork_end``, and ``sample_size`` together.

    Args:
        db: Active :class:`~db.Database` connection.
        pollster_id: Primary key of the pollster in the database.
        map_id: Primary key of the constituency map in the database.
        parsed: :class:`ParsedPoll` whose fieldwork and sample metadata are
            used for the lookup.

    Returns:
        The matching :class:`~models.Poll` row, or ``None`` if no match exists.
    """
    with db.session() as session:
        return session.execute(
            select(Poll).where(
                Poll.pollster_id == pollster_id,
                Poll.map_id == map_id,
                Poll.fieldwork_start == parsed.fieldwork_start,
                Poll.fieldwork_end == parsed.fieldwork_end,
                Poll.sample_size == parsed.sample_size,
            )
        ).scalar_one_or_none()


def build_import_plan(
    db: Database,
    *,
    xlsx_url: str = DEFAULT_XLSX_URL,
    map_name: str = DEFAULT_MAP_NAME,
    pollster_identifier: str = DEFAULT_POLLSTER_IDENTIFIER,
    year_hint: int | None = None,
) -> ImportPlan:
    """Download and parse a BMG Research XLSX file and build a dry-run import plan.

    Fetches the workbook, parses poll metadata and voting-intention figures,
    resolves all party and region IDs from the database, and checks whether the
    pollster and poll already exist.  No data is written to the database.

    Args:
        db: Active :class:`~db.Database` connection.
        xlsx_url: URL of the BMG Research XLSX file to import.  Defaults to
            :data:`DEFAULT_XLSX_URL`.
        map_name: Name of the constituency map to associate the poll with.
            Must match a map already present in the database.  Defaults to
            :data:`DEFAULT_MAP_NAME`.
        pollster_identifier: Identifier string for the pollster record.
            Defaults to :data:`DEFAULT_POLLSTER_IDENTIFIER`.
        year_hint: Optional fallback year (four-digit integer) used when the
            URL and fieldwork strings contain no explicit year.

    Returns:
        A fully populated :class:`ImportPlan` describing what would be created
        or updated in the database when :func:`commit_import_plan` is called.

    Raises:
        ValueError: If the specified map is not found in the database, if the
            workbook cannot be fetched or parsed, or if any required parties
            are missing from the database.
    """
    poll_map = db.get_map_by_name(map_name)
    if poll_map is None:
        raise ValueError(f"Map not found: {map_name!r}")

    workbook = extract_workbook(xlsx_url)
    parsed = parse_poll(workbook, source_url=xlsx_url, year_hint=year_hint)

    pollster = db.get_pollster_by_identifier(pollster_identifier)
    pollster_exists = pollster is not None

    party_by_name = {party.name: party for party in db.get_all_parties()}
    missing_parties = [
        party_name
        for party_name in sorted(set(PARTY_NAME_MAP.values()))
        if party_name not in party_by_name
    ]
    if missing_parties:
        raise ValueError(
            "Missing parties in database (run party importer first): "
            f"{missing_parties}"
        )

    rows: list[PlannedPollRow] = []
    regions = db.get_regions_for_map(poll_map.id)
    region_ids_by_name: dict[str, int] = {region.name: region.id for region in regions}

    for party_name, region_values in parsed.party_region_percentages.items():
        national_percentage = region_values.get(NATIONAL_KEY, 0.0)
        rows.append(
            PlannedPollRow(
                party_id=party_by_name[party_name].id,
                party_name=party_name,
                region_id=None,
                region_name="National",
                percentage=national_percentage,
            )
        )

        for region_name, region_id in region_ids_by_name.items():
            rows.append(
                PlannedPollRow(
                    party_id=party_by_name[party_name].id,
                    party_name=party_name,
                    region_id=region_id,
                    region_name=region_name,
                    percentage=region_values.get(region_name, 0.0),
                )
            )

    existing_poll = (
        _find_existing_poll(db, pollster.id, poll_map.id, parsed)
        if pollster is not None
        else None
    )

    return ImportPlan(
        pollster_identifier=pollster_identifier,
        pollster_name=(pollster.name if pollster else "BMG Research"),
        pollster_id=(pollster.id if pollster else None),
        pollster_exists=pollster_exists,
        regions_mapping="",
        map_id=poll_map.id,
        map_name=poll_map.name,
        source_url=xlsx_url,
        parsed=parsed,
        poll_id=(existing_poll.id if existing_poll else None),
        poll_exists=existing_poll is not None,
        rows=rows,
    )


def commit_import_plan(
    db: Database,
    plan: ImportPlan,
    *,
    replace_rows: bool = False,
) -> PollImportResult:
    """Write an :class:`ImportPlan` to the database, creating or updating records as needed.

    Execution order:

    1. Creates the pollster if it does not yet exist.
    2. Creates the poll if it does not yet exist; otherwise updates
       ``source_url`` on the existing poll if it has changed.
    3. If the poll already has rows and ``replace_rows`` is ``False``, returns
       immediately without touching rows (``skipped_existing_rows=True``).
    4. If ``replace_rows`` is ``True``, deletes existing rows before inserting.
    5. Inserts all :class:`PlannedPollRow` entries from the plan.

    Args:
        db: Active :class:`~db.Database` connection.
        plan: The :class:`ImportPlan` produced by :func:`build_import_plan`.
        replace_rows: If ``True``, delete any existing :class:`~models.PollRow`
            records for the poll before inserting new ones.  Defaults to
            ``False``, in which case existing rows are left untouched and the
            result will have ``skipped_existing_rows=True``.

    Returns:
        A :class:`~polls.import_types.PollImportResult` summarising what was
        created, replaced, inserted, or skipped.

    Raises:
        ValueError: If the pollster lookup fails unexpectedly during the commit
            phase.
    """
    if plan.pollster_exists:
        pollster = db.get_pollster_by_identifier(plan.pollster_identifier)
        if pollster is None:
            raise ValueError("Pollster lookup failed during commit")
        pollster_id = pollster.id
        created_pollster = False
    else:
        created = db.add_pollster(
            name=plan.pollster_name,
            identifier=plan.pollster_identifier,
            weight=1.0,
            regions_mapping=plan.regions_mapping,
        )
        pollster_id = created.id
        created_pollster = True

    existing_poll = _find_existing_poll(db, pollster_id, plan.map_id, plan.parsed)
    if existing_poll is None:
        created_poll = db.add_poll(
            pollster_id=pollster_id,
            map_id=plan.map_id,
            fieldwork_start=plan.parsed.fieldwork_start,
            fieldwork_end=plan.parsed.fieldwork_end,
            sample_size=plan.parsed.sample_size,
            source_url=plan.source_url,
        )
        poll_id = created_poll.id
        created_poll_row = True
    else:
        with db.session() as session:
            db_poll = session.get(Poll, existing_poll.id)
            if db_poll is not None and db_poll.source_url != plan.source_url:
                db_poll.source_url = plan.source_url
        poll_id = existing_poll.id
        created_poll_row = False

    with db.session() as session:
        existing_rows = session.execute(
            select(PollRow).where(PollRow.poll_id == poll_id)
        ).scalars().all()

        if existing_rows and not replace_rows:
            return PollImportResult(
                created_pollster=created_pollster,
                created_poll=created_poll_row,
                poll_id=poll_id,
                inserted_rows=0,
                replaced_rows=0,
                skipped_existing_rows=True,
            )

        replaced_rows = 0
        if existing_rows and replace_rows:
            for existing_row in existing_rows:
                session.delete(existing_row)
                replaced_rows += 1

        for row in plan.rows:
            session.add(
                PollRow(
                    poll_id=poll_id,
                    region_id=row.region_id,
                    party_id=row.party_id,
                    percentage=row.percentage,
                )
            )

    return PollImportResult(
        created_pollster=created_pollster,
        created_poll=created_poll_row,
        poll_id=poll_id,
        inserted_rows=len(plan.rows),
        replaced_rows=replaced_rows,
        skipped_existing_rows=False,
    )


def _cli_preview(plan: ImportPlan) -> None:
    """Print a human-readable dry-run summary of an :class:`ImportPlan` to stdout.

    Displays the parsed fieldwork dates, sample size, pollster status, poll
    status, and a line for each planned poll row.  No data is written to the
    database.

    Args:
        plan: The :class:`ImportPlan` to preview.
    """
    print(
        "Parsed poll: "
        f"fieldwork={plan.parsed.fieldwork_start} to {plan.parsed.fieldwork_end}, "
        f"sample={plan.parsed.sample_size}"
    )
    if plan.pollster_exists:
        print(f"pollster exists: {plan.pollster_identifier}")
    else:
        print(f"[dry-run] would create pollster: {plan.pollster_identifier}")

    if plan.poll_exists and plan.poll_id is not None:
        print(f"poll exists: {plan.poll_id}")
    else:
        print("[dry-run] would create poll")

    for row in plan.rows:
        print(
            "[dry-run] would insert row: "
            f"party={row.party_name}, region={row.region_name}, "
            f"region_id={row.region_id}, pct={row.percentage:.2f}"
        )


def main() -> None:
    """CLI entry point for importing a BMG Research poll from an XLSX URL.

    Parses command-line arguments, builds an import plan, and either prints a
    dry-run preview or commits the plan to the database.

    Command-line arguments:

    * ``--xlsx-url`` (str, optional): URL of the XLSX file to import.
      Defaults to :data:`DEFAULT_XLSX_URL`.
    * ``--map-name`` (str, optional): Constituency map name.
      Defaults to :data:`DEFAULT_MAP_NAME`.
    * ``--pollster-identifier`` (str, optional): Pollster identifier.
      Defaults to :data:`DEFAULT_POLLSTER_IDENTIFIER`.
    * ``--year-hint`` (int, optional): Fallback year when none is found in the
      URL or fieldwork string.
    * ``--replace-rows`` (flag): Delete and re-insert existing poll rows.
    * ``--dry-run`` (flag): Print a preview without writing to the database.
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx-url", default=DEFAULT_XLSX_URL)
    parser.add_argument("--map-name", default=DEFAULT_MAP_NAME)
    parser.add_argument("--pollster-identifier", default=DEFAULT_POLLSTER_IDENTIFIER)
    parser.add_argument("--year-hint", type=int, default=None)
    parser.add_argument(
        "--replace-rows",
        action="store_true",
        help="Delete existing rows for the poll before inserting new ones",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    db = Database()
    print(f"Fetching XLSX: {args.xlsx_url}")

    plan = build_import_plan(
        db,
        xlsx_url=args.xlsx_url,
        map_name=args.map_name,
        pollster_identifier=args.pollster_identifier,
        year_hint=args.year_hint,
    )

    if args.dry_run:
        _cli_preview(plan)
        return

    result = commit_import_plan(db, plan, replace_rows=args.replace_rows)
    if result.created_pollster:
        print(f"created pollster: {args.pollster_identifier}")
    else:
        print(f"pollster exists: {args.pollster_identifier}")

    if result.created_poll:
        print(f"created poll: {result.poll_id}")
    else:
        print(f"poll exists: {result.poll_id}")

    if result.skipped_existing_rows:
        print(
            f"poll {result.poll_id} already has rows; "
            "use --replace-rows to overwrite"
        )
    else:
        if result.replaced_rows:
            print(f"deleted existing rows: {result.replaced_rows}")
        print(f"inserted poll rows: {result.inserted_rows}")


if __name__ == "__main__":
    main()
