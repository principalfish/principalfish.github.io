#!/usr/bin/env python3
"""Import a Find Out Now poll directly from an XLSX URL."""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.request import urlopen

from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import select

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from db import Database
from models import Poll, PollRow, Pollster
from polls.import_types import PollImportResult

DEFAULT_XLSX_URL = "https://cms.findoutnow.co.uk/app/uploads/2026/02/11th-February-VI-Find-Out-Now.xlsx"
DEFAULT_MAP_NAME = "UK Constituencies post 2022"
DEFAULT_POLLSTER_IDENTIFIER = "find_out_now"

PARTY_NAME_MAP = {
    "Conservative": "Conservative",
    "Labour": "Labour",
    "Liberal Democrat": "Liberal Democrats",
    "Liberal Democrats": "Liberal Democrats",
    "Reform UK": "Reform UK",
    "Green": "Green",
    "Green Party": "Green",
    "SNP": "Scottish National Party",
    "Scottish National Party": "Scottish National Party",
    "Scottish National Party (SNP)": "Scottish National Party",
    "Plaid Cymru": "Plaid Cymru",
    "Other": "Other",
}

REGION_HEADER_TO_INTERNAL = {
    "East Midlands": "East Midlands",
    "East of England": "East of England",
    "London": "London",
    "North East": "North East England",
    "North West": "North West England",
    "Scotland": "Scotland",
    "South East": "South East England",
    "South West": "South West England",
    "Wales": "Wales",
    "West Midlands": "West Midlands",
    "Yorkshire and the Humber": "Yorkshire and The Humber",
}

NATIONAL_KEY = "__national__"

class ParsedPoll(BaseModel):
    """Parsed poll data extracted from source."""

    sample_size: int = Field(gt=0)
    fieldwork_start: date
    fieldwork_end: date
    party_region_percentages: dict[str, dict[str, float]]


def _maybe_adjust_fieldwork_year_from_url(parsed: ParsedPoll, url_year: int | None) -> ParsedPoll:
    """Correct a fieldwork year that appears to be off by one due to a year-boundary edge case.

    When the URL contains a year and the parsed fieldwork dates fall in the final
    months of the preceding year (month > 3 is excluded), the dates are likely
    in the URL year rather than the parsed year.  Both start and end are adjusted
    together only when they share the same year and that year is exactly one less
    than the URL year.

    Args:
        parsed: The poll data as extracted from the workbook.
        url_year: Four-digit year inferred from the XLSX URL, or None if no year
            was found in the URL.

    Returns:
        A new ParsedPoll with corrected fieldwork dates, or the original ParsedPoll
        unchanged if no adjustment is warranted.
    """
    if url_year is None:
        return parsed

    start = parsed.fieldwork_start
    end = parsed.fieldwork_end

    if start.year != end.year:
        return parsed
    if end.year != (url_year - 1):
        return parsed
    if end.month > 3:
        return parsed

    try:
        adjusted_start = start.replace(year=url_year)
        adjusted_end = end.replace(year=url_year)
    except ValueError:
        return parsed

    return ParsedPoll(
        sample_size=parsed.sample_size,
        fieldwork_start=adjusted_start,
        fieldwork_end=adjusted_end,
        party_region_percentages=parsed.party_region_percentages,
    )


class PlannedPollRow(BaseModel):
    """A single poll row planned for DB insertion."""

    party_id: int
    party_name: str
    region_id: int | None
    region_name: str
    percentage: float = Field(ge=0, le=100)


class ImportPlan(BaseModel):
    """Full import plan: pollster, poll metadata, and rows."""

    pollster_identifier: str
    pollster_name: str
    pollster_id: int | None
    pollster_exists: bool
    regions_mapping: str
    map_id: int
    map_name: str
    source_url: str
    parsed: ParsedPoll
    poll_id: int | None
    poll_exists: bool
    rows: list[PlannedPollRow]


def normalize_name(value: str) -> str:
    """Normalise a name by collapsing whitespace and lowercasing.

    Args:
        value: Raw string to normalise.

    Returns:
        Lowercase string with leading/trailing whitespace stripped and internal
        runs of whitespace collapsed to a single space.
    """
    return " ".join(value.strip().split()).lower()


def _month_number(month_text: str) -> int | None:
    """Convert a month name to its numeric equivalent (1–12).

    Args:
        month_text: Full English month name, case-insensitive
            (e.g. ``"January"``, ``"february"``).

    Returns:
        Integer month number (1 for January … 12 for December), or ``None`` if
        the name is not recognised.
    """
    month_map = {
        "january": 1,
        "february": 2,
        "march": 3,
        "april": 4,
        "may": 5,
        "june": 6,
        "july": 7,
        "august": 8,
        "september": 9,
        "october": 10,
        "november": 11,
        "december": 12,
    }
    return month_map.get(month_text.lower())


def parse_fieldwork(value: str, *, year_hint: int | None = None) -> tuple[date, date]:
    """Parse a free-text fieldwork date string into a (start, end) date pair.

    Handles several formats found in Find Out Now workbooks:

    - Cross-month range with year: ``"28th January - 3rd February 2026"``
    - Same-month range with year: ``"1-5 February 2026"``
    - Single day with year: ``"5th February 2026"``
    - Same-month range without year: ``"1-5 February"`` (requires ``year_hint``)
    - Single day without year: ``"5th February"`` (requires ``year_hint``)

    For cross-month ranges the start year is inferred as one year earlier than
    the end year when the start month is numerically greater than the end month
    (i.e. the range straddles a year boundary).

    Args:
        value: Raw fieldwork date string extracted from the workbook.
        year_hint: Four-digit year to use when the string contains no explicit
            year.  Required for the no-year patterns; ignored otherwise.

    Returns:
        A ``(fieldwork_start, fieldwork_end)`` tuple of :class:`datetime.date`
        objects.  For a single-day poll both values are the same date.

    Raises:
        ValueError: If the string does not match any recognised pattern, if a
            month name cannot be resolved, or if ``year_hint`` is ``None`` when
            required by a no-year pattern.
    """
    normalized = re.sub(r"\s+", " ", value.strip().replace("–", "-"))

    cross_month_pattern = re.compile(
        r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\s*-\s*(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\s+(\d{4})"
    )
    range_same_month_pattern = re.compile(
        r"(\d{1,2})(?:st|nd|rd|th)?\s*-\s*(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\s+(\d{4})"
    )
    single_day_pattern = re.compile(
        r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\s+(\d{4})"
    )
    range_same_month_no_year_pattern = re.compile(
        r"(\d{1,2})(?:st|nd|rd|th)?\s*-\s*(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)"
    )
    single_day_no_year_pattern = re.compile(
        r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)"
    )

    match = cross_month_pattern.search(normalized)
    if match:
        day_start = int(match.group(1))
        month_start = _month_number(match.group(2))
        day_end = int(match.group(3))
        month_end = _month_number(match.group(4))
        year_end = int(match.group(5))
        if month_start is None or month_end is None:
            raise ValueError(f"Could not parse fieldwork string: {value!r}")
        year_start = year_end - 1 if month_start > month_end else year_end
        return date(year_start, month_start, day_start), date(year_end, month_end, day_end)

    match = range_same_month_pattern.search(normalized)
    if match:
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        year = int(match.group(4))
        if month is None:
            raise ValueError(f"Could not parse fieldwork string: {value!r}")
        return date(year, month, day_start), date(year, month, day_end)

    match = single_day_pattern.search(normalized)
    if match:
        day = int(match.group(1))
        month = _month_number(match.group(2))
        year = int(match.group(3))
        if month is None:
            raise ValueError(f"Could not parse fieldwork string: {value!r}")
        parsed_day = date(year, month, day)
        return parsed_day, parsed_day

    match = range_same_month_no_year_pattern.search(normalized)
    if match:
        if year_hint is None:
            raise ValueError(f"Could not parse fieldwork year from string: {value!r}")
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        if month is None:
            raise ValueError(f"Could not parse fieldwork string: {value!r}")
        return date(year_hint, month, day_start), date(year_hint, month, day_end)

    match = single_day_no_year_pattern.search(normalized)
    if match:
        if year_hint is None:
            raise ValueError(f"Could not parse fieldwork year from string: {value!r}")
        day = int(match.group(1))
        month = _month_number(match.group(2))
        if month is None:
            raise ValueError(f"Could not parse fieldwork string: {value!r}")
        parsed_day = date(year_hint, month, day)
        return parsed_day, parsed_day

    raise ValueError(f"Could not parse fieldwork string: {value!r}")


def _infer_year_hint_from_url(url: str) -> int | None:
    """Extract a four-digit year from a URL string.

    Searches for the first occurrence of a year in the range 2000–2099 within
    the URL path.

    Args:
        url: Full URL string, e.g. the XLSX download URL.

    Returns:
        The first matching four-digit year as an integer, or ``None`` if no
        year matching ``20XX`` is found in the URL.
    """
    match = re.search(r"(20\d{2})", url)
    if not match:
        return None
    return int(match.group(1))


def extract_workbook(xlsx_url: str) -> Any:
    """Download an XLSX file from a URL and open it as an openpyxl workbook.

    Args:
        xlsx_url: Fully-qualified URL pointing to the XLSX file to fetch.

    Returns:
        An openpyxl ``Workbook`` object opened with ``data_only=True`` so that
        cell values are read rather than formulas.
    """
    payload = urlopen(xlsx_url).read()
    return load_workbook(filename=BytesIO(payload), data_only=True)


def _cell_text(value: object) -> str:
    """Convert a raw cell value to a stripped string.

    Args:
        value: Raw value from an openpyxl cell (may be ``None``, a number, a
            string, etc.).

    Returns:
        String representation of the value with leading/trailing whitespace
        removed, or an empty string if ``value`` is ``None``.
    """
    if value is None:
        return ""
    return str(value).strip()


def _find_label_value(ws: Any, label_fragment: str) -> str | None:
    """Search a worksheet for a label and return the first non-empty value to its right.

    Scans the top-left region of the worksheet (rows 1–34, columns 1–6) looking
    for a cell whose text contains ``label_fragment`` (case-insensitive).  Once
    found, checks up to three cells to the right of that label cell and returns
    the first non-empty value encountered.

    Args:
        ws: An openpyxl ``Worksheet`` object to search.
        label_fragment: Substring to look for in cell text, matched
            case-insensitively (e.g. ``"Fieldwork date"``).

    Returns:
        The stripped string value of the first non-empty cell to the right of
        the matching label, or ``None`` if no matching label or adjacent value
        is found.
    """
    needle = label_fragment.lower()
    for row in range(1, 35):
        for col in range(1, 7):
            label = _cell_text(ws.cell(row, col).value)
            if not label:
                continue
            if needle in label.lower():
                for offset in range(1, 4):
                    candidate = _cell_text(ws.cell(row, col + offset).value)
                    if candidate:
                        return candidate
    return None


def _as_int(value: str) -> int:
    """Extract the first contiguous run of digits from a string and return it as an integer.

    Strips all non-digit characters (e.g. commas, spaces) before parsing, which
    handles formatted numbers such as ``"1,234"`` or ``"n=1234"``.

    Args:
        value: Raw string expected to contain at least one digit
            (e.g. a sample-size cell value like ``"1,234"``).

    Returns:
        Integer formed from all digit characters found in ``value``.

    Raises:
        ValueError: If ``value`` contains no digit characters.
    """
    digits = re.sub(r"[^0-9]", "", value)
    if not digits:
        raise ValueError(f"Could not parse sample size from value: {value!r}")
    return int(digits)


def _to_percentage(value: object) -> float:
    """Convert a cell value to a rounded integer-valued percentage float.

    Accepts both decimal proportions (0.0–1.0) and already-scaled percentages
    (values outside 0–1 are assumed to already be in percent).  The result is
    rounded to the nearest integer and returned as a float.

    Args:
        value: Raw cell value to convert.  Must be coercible to ``float``
            (e.g. an openpyxl numeric cell value).  Must not be ``None``.

    Returns:
        Percentage value as a float with no fractional part (e.g. ``42.0``).

    Raises:
        ValueError: If ``value`` is ``None``.
    """
    if value is None:
        raise ValueError("Encountered empty percentage cell")
    number = float(value)  # type: ignore[arg-type]
    pct = number * 100.0 if 0.0 <= number <= 1.0 else number
    return float(int(round(pct)))


def parse_poll(workbook: Any, *, fieldwork_year_hint: int | None = None) -> ParsedPoll:
    """Extract poll metadata and vote-intention percentages from an openpyxl workbook.

    Reads the cover sheet for fieldwork dates and sample size, then attempts to
    read regional vote-intention data from a ``"Headline VI"`` sheet (preferred)
    or falls back to a ``"Q2"`` sheet that contains national-only figures.

    Party labels are normalised via ``PARTY_NAME_MAP``; region headers are
    mapped to internal names via ``REGION_HEADER_TO_INTERNAL``.  Rows labelled
    ``"filtered n"`` (case-insensitive) act as a sentinel stopping row iteration.

    Args:
        workbook: An openpyxl ``Workbook`` object, typically produced by
            :func:`extract_workbook`.
        fieldwork_year_hint: Four-digit year to supply to :func:`parse_fieldwork`
            when the fieldwork date string contains no explicit year.  Optional;
            defaults to ``None``.

    Returns:
        A :class:`ParsedPoll` containing the sample size, fieldwork date range,
        and a nested mapping of ``{canonical_party: {region_key: percentage}}``.
        The special key ``NATIONAL_KEY`` (``"__national__"``) holds the national
        figure for each party.

    Raises:
        ValueError: If the fieldwork date or sample size cannot be found or
            parsed, if the regional header row or ``"All"`` column cannot be
            located, if fewer than six region columns are resolved, if neither a
            ``"Headline VI"`` nor a ``"Q2"`` sheet exists, or if any of the
            required party rows are absent from the parsed data.
    """
    cover_sheet_name = "Cover page" if "Cover page" in workbook.sheetnames else workbook.sheetnames[0]
    cover = workbook[cover_sheet_name]

    fieldwork_raw = _find_label_value(cover, "Fieldwork date") or _cell_text(cover["C5"].value)
    if not fieldwork_raw:
        raise ValueError("Fieldwork date not found in workbook")
    fieldwork_start, fieldwork_end = parse_fieldwork(fieldwork_raw, year_hint=fieldwork_year_hint)

    sample_raw = _find_label_value(cover, "Sample size") or _cell_text(cover["C6"].value)
    if not sample_raw:
        raise ValueError("Sample size not found in workbook")
    sample_size = _as_int(sample_raw)

    headline_sheet = "Headline VI" if "Headline VI" in workbook.sheetnames else None
    if headline_sheet is None:
        for sheet_name in workbook.sheetnames:
            if "headline" in sheet_name.lower():
                headline_sheet = sheet_name
                break
    party_region_percentages: dict[str, dict[str, float]] = {}

    if headline_sheet is not None:
        sheet = workbook[headline_sheet]

        header_row = None
        for row in range(1, 30):
            values = [_cell_text(sheet.cell(row, col).value) for col in range(1, 50)]
            if "All" in values and "East Midlands" in values:
                header_row = row
                break
        if header_row is None:
            raise ValueError("Could not locate regional header row in Headline VI sheet")

        region_columns: dict[int, str] = {}
        national_column: int | None = None
        for col in range(1, 80):
            header = _cell_text(sheet.cell(header_row, col).value)
            if header == "All":
                national_column = col
            if header in REGION_HEADER_TO_INTERNAL:
                region_columns[col] = REGION_HEADER_TO_INTERNAL[header]
        if national_column is None:
            raise ValueError("Could not locate 'All' (national) column in Headline VI")
        if len(region_columns) < 6:
            raise ValueError("Could not resolve sufficient region columns in Headline VI")

        for row in range(header_row + 1, header_row + 40):
            party_label = _cell_text(sheet.cell(row, 1).value)
            if not party_label:
                continue
            if normalize_name(party_label) == "filtered n":
                break
            canonical_party = PARTY_NAME_MAP.get(party_label)
            if canonical_party is None:
                continue

            region_values: dict[str, float] = {}
            region_values[NATIONAL_KEY] = _to_percentage(sheet.cell(row, national_column).value)
            for col, internal_region in region_columns.items():
                region_values[internal_region] = _to_percentage(sheet.cell(row, col).value)
            party_region_percentages[canonical_party] = region_values
    else:
        q2_sheet_name = "Q2" if "Q2" in workbook.sheetnames else None
        if q2_sheet_name is None:
            for sheet_name in workbook.sheetnames:
                if sheet_name.lower().startswith("q2"):
                    q2_sheet_name = sheet_name
                    break
        if q2_sheet_name is None:
            raise ValueError("Could not find 'Headline VI' or 'Q2' sheet in workbook")

        sheet = workbook[q2_sheet_name]
        header_row = None
        national_column = None
        for row in range(1, 30):
            values = [_cell_text(sheet.cell(row, col).value) for col in range(1, 30)]
            if "All" in values:
                header_row = row
                national_column = values.index("All") + 1
                break
        if header_row is None or national_column is None:
            raise ValueError("Could not locate 'All' (national) column in Q2 sheet")

        for row in range(header_row + 1, header_row + 50):
            party_label = _cell_text(sheet.cell(row, 1).value)
            if not party_label:
                continue
            if normalize_name(party_label) == "filtered n":
                break
            canonical_party = PARTY_NAME_MAP.get(party_label)
            if canonical_party is None:
                continue
            party_region_percentages[canonical_party] = {
                NATIONAL_KEY: _to_percentage(sheet.cell(row, national_column).value)
            }

    required_parties = {
        "Conservative",
        "Labour",
        "Liberal Democrats",
        "Reform UK",
        "Green",
        "Scottish National Party",
        "Plaid Cymru",
        "Other",
    }
    missing = sorted(required_parties - set(party_region_percentages.keys()))
    if missing:
        raise ValueError(f"Missing expected party rows in workbook: {missing}")

    return ParsedPoll(
        sample_size=sample_size,
        fieldwork_start=fieldwork_start,
        fieldwork_end=fieldwork_end,
        party_region_percentages=party_region_percentages,
    )


def _find_existing_poll(
    db: Database,
    pollster_id: int,
    map_id: int,
    parsed: ParsedPoll,
) -> Poll | None:
    """Query the database for a poll matching the given metadata.

    Looks up a poll by pollster, map, fieldwork date range, and sample size.
    All five fields must match for a poll to be considered a duplicate.

    Args:
        db: Active :class:`Database` instance used to open the session.
        pollster_id: Primary key of the pollster to match.
        map_id: Primary key of the constituency map to match.
        parsed: Parsed poll data providing the fieldwork dates and sample size
            to match against.

    Returns:
        The matching :class:`Poll` ORM object if one exists, or ``None``.
    """
    with db.session() as session:
        return session.execute(
            select(Poll).where(
                Poll.pollster_id == pollster_id,
                Poll.map_id == map_id,
                Poll.fieldwork_start == parsed.fieldwork_start,
                Poll.fieldwork_end == parsed.fieldwork_end,
                Poll.sample_size == parsed.sample_size,
            )
        ).scalar_one_or_none()


def build_import_plan(
    db: Database,
    *,
    xlsx_url: str = DEFAULT_XLSX_URL,
    map_name: str = DEFAULT_MAP_NAME,
    pollster_identifier: str = DEFAULT_POLLSTER_IDENTIFIER,
    fieldwork_year_hint: int | None = None,
) -> ImportPlan:
    """Build a complete import plan without writing anything to the database.

    Downloads and parses the XLSX workbook, resolves the constituency map,
    regions, pollster, and parties from the database, and assembles the full
    list of :class:`PlannedPollRow` objects that would be inserted.  No data
    is written; the returned plan is suitable for preview (dry-run) or for
    passing directly to :func:`commit_import_plan`.

    If ``fieldwork_year_hint`` is not provided, a year is inferred from the
    XLSX URL.  After parsing, a year-boundary adjustment is applied via
    :func:`_maybe_adjust_fieldwork_year_from_url`.

    Args:
        db: Active :class:`Database` instance used for all lookups.
        xlsx_url: Fully-qualified URL of the XLSX file to fetch.  Defaults to
            ``DEFAULT_XLSX_URL``.
        map_name: Name of the constituency map to look up in the database.
            Defaults to ``DEFAULT_MAP_NAME``.
        pollster_identifier: Unique string identifier for the pollster record.
            Defaults to ``DEFAULT_POLLSTER_IDENTIFIER``.
        fieldwork_year_hint: Optional four-digit year to use when the fieldwork
            date string in the workbook contains no explicit year.  When
            ``None``, the year is inferred from ``xlsx_url``.

    Returns:
        A fully-populated :class:`ImportPlan` describing what would be created
        or updated in the database.

    Raises:
        ValueError: If the named map is not found in the database, if any
            required party is absent from the database, or if workbook parsing
            fails (see :func:`parse_poll`).
    """
    poll_map = db.get_map_by_name(map_name)
    if poll_map is None:
        raise ValueError(f"Map not found: {map_name!r}")

    regions = db.get_regions_for_map(poll_map.id)
    regions_by_name = {normalize_name(region.name): region for region in regions}

    region_ids_by_name: dict[str, int] = {}
    for region in regions:
        region_ids_by_name[region.name] = region.id

    regions_mapping = "\n".join(
        f"{region_name}:{region_ids_by_name[region_name]}"
        for region_name in sorted(region_ids_by_name.keys())
    )

    workbook = extract_workbook(xlsx_url)
    effective_year_hint = fieldwork_year_hint
    if effective_year_hint is None:
        effective_year_hint = _infer_year_hint_from_url(xlsx_url)
    parsed = parse_poll(workbook, fieldwork_year_hint=effective_year_hint)
    parsed = _maybe_adjust_fieldwork_year_from_url(parsed, effective_year_hint)

    pollster = db.get_pollster_by_identifier(pollster_identifier)
    pollster_exists = pollster is not None

    party_by_name = {party.name: party for party in db.get_all_parties()}
    missing_parties = [
        party_name
        for party_name in sorted(set(PARTY_NAME_MAP.values()))
        if party_name not in party_by_name
    ]
    if missing_parties:
        raise ValueError(
            "Missing parties in database (run party importer first): "
            f"{missing_parties}"
        )

    rows: list[PlannedPollRow] = []
    for party_name, region_values in parsed.party_region_percentages.items():
        party_id = party_by_name[party_name].id
        national_percentage = region_values.get(NATIONAL_KEY)
        if national_percentage is not None:
            rows.append(
                PlannedPollRow(
                    party_id=party_id,
                    party_name=party_name,
                    region_id=None,
                    region_name="National",
                    percentage=national_percentage,
                )
            )

        for region_name, region_id in region_ids_by_name.items():
            percentage = region_values.get(region_name, 0.0)
            rows.append(
                PlannedPollRow(
                    party_id=party_id,
                    party_name=party_name,
                    region_id=region_id,
                    region_name=region_name,
                    percentage=percentage,
                )
            )

    existing_poll = (
        _find_existing_poll(db, pollster.id, poll_map.id, parsed)
        if pollster is not None
        else None
    )

    return ImportPlan(
        pollster_identifier=pollster_identifier,
        pollster_name=(pollster.name if pollster else "Find Out Now"),
        pollster_id=(pollster.id if pollster else None),
        pollster_exists=pollster_exists,
        regions_mapping=regions_mapping,
        map_id=poll_map.id,
        map_name=poll_map.name,
        source_url=xlsx_url,
        parsed=parsed,
        poll_id=(existing_poll.id if existing_poll else None),
        poll_exists=existing_poll is not None,
        rows=rows,
    )


def commit_import_plan(
    db: Database,
    plan: ImportPlan,
    *,
    replace_rows: bool = False,
) -> PollImportResult:
    """Execute an import plan, writing pollster, poll, and row records to the database.

    If the pollster does not yet exist it is created; if it does, its
    ``regions_mapping`` is updated if it has changed.  The same logic applies
    to the poll record.  For poll rows, the default behaviour is to skip
    insertion when rows already exist; passing ``replace_rows=True`` deletes
    the existing rows before inserting the new set.

    Args:
        db: Active :class:`Database` instance used for all writes.
        plan: A fully-populated :class:`ImportPlan` as returned by
            :func:`build_import_plan`.
        replace_rows: When ``True``, any existing :class:`PollRow` records for
            the poll are deleted before the new rows are inserted.  When
            ``False`` (default), existing rows are left unchanged and insertion
            is skipped entirely.

    Returns:
        A :class:`PollImportResult` summarising what was created, replaced, or
        skipped, including ``poll_id``, row counts, and boolean flags for
        ``created_pollster``, ``created_poll``, and ``skipped_existing_rows``.

    Raises:
        ValueError: If the pollster lookup fails during commit (indicates an
            unexpected database inconsistency).
    """
    if plan.pollster_exists:
        pollster = db.get_pollster_by_identifier(plan.pollster_identifier)
        if pollster is None:
            raise ValueError("Pollster lookup failed during commit")
        with db.session() as session:
            db_pollster = session.get(Pollster, pollster.id)
            if db_pollster is not None and db_pollster.regions_mapping != plan.regions_mapping:
                db_pollster.regions_mapping = plan.regions_mapping
        pollster_id = pollster.id
        created_pollster = False
    else:
        created = db.add_pollster(
            name=plan.pollster_name,
            identifier=plan.pollster_identifier,
            weight=1.0,
            regions_mapping=plan.regions_mapping,
        )
        pollster_id = created.id
        created_pollster = True

    existing_poll = _find_existing_poll(db, pollster_id, plan.map_id, plan.parsed)
    if existing_poll is None:
        created_poll = db.add_poll(
            pollster_id=pollster_id,
            map_id=plan.map_id,
            fieldwork_start=plan.parsed.fieldwork_start,
            fieldwork_end=plan.parsed.fieldwork_end,
            sample_size=plan.parsed.sample_size,
            source_url=plan.source_url,
        )
        poll_id = created_poll.id
        created_poll_row = True
    else:
        with db.session() as session:
            db_poll = session.get(Poll, existing_poll.id)
            if db_poll is not None and db_poll.source_url != plan.source_url:
                db_poll.source_url = plan.source_url
        poll_id = existing_poll.id
        created_poll_row = False

    with db.session() as session:
        existing_rows = session.execute(
            select(PollRow).where(PollRow.poll_id == poll_id)
        ).scalars().all()

        if existing_rows and not replace_rows:
            return PollImportResult(
                created_pollster=created_pollster,
                created_poll=created_poll_row,
                poll_id=poll_id,
                inserted_rows=0,
                replaced_rows=0,
                skipped_existing_rows=True,
            )

        replaced_rows = 0
        if existing_rows and replace_rows:
            for existing_row in existing_rows:
                session.delete(existing_row)
                replaced_rows += 1

        for row in plan.rows:
            session.add(
                PollRow(
                    poll_id=poll_id,
                    region_id=row.region_id,
                    party_id=row.party_id,
                    percentage=row.percentage,
                )
            )

    return PollImportResult(
        created_pollster=created_pollster,
        created_poll=created_poll_row,
        poll_id=poll_id,
        inserted_rows=len(plan.rows),
        replaced_rows=replaced_rows,
        skipped_existing_rows=False,
    )


def _cli_preview(plan: ImportPlan) -> None:
    """Print a human-readable dry-run summary of an import plan to stdout.

    Displays the parsed fieldwork dates and sample size, whether the pollster
    and poll already exist, and one line per planned poll row showing party,
    region, region ID, and percentage.

    Args:
        plan: A fully-populated :class:`ImportPlan` as returned by
            :func:`build_import_plan`.
    """
    print(
        "Parsed poll: "
        f"fieldwork={plan.parsed.fieldwork_start} to {plan.parsed.fieldwork_end}, "
        f"sample={plan.parsed.sample_size}"
    )
    if plan.pollster_exists:
        print(f"pollster exists: {plan.pollster_identifier}")
    else:
        print(f"[dry-run] would create pollster: {plan.pollster_identifier}")

    if plan.poll_exists and plan.poll_id is not None:
        print(f"poll exists: {plan.poll_id}")
    else:
        print("[dry-run] would create poll")

    for row in plan.rows:
        print(
            "[dry-run] would insert row: "
            f"party={row.party_name}, region={row.region_name}, "
            f"region_id={row.region_id}, pct={row.percentage:.2f}"
        )


def main() -> None:
    """Entry point for the Find Out Now poll importer CLI.

    Parses command-line arguments, fetches and parses the XLSX workbook, builds
    an import plan, and either previews it (``--dry-run``) or commits it to the
    database.  Prints a status summary to stdout after each step.

    CLI arguments:
        --xlsx-url (str): URL of the XLSX file to import.  Defaults to
            ``DEFAULT_XLSX_URL``.
        --map-name (str): Name of the constituency map to use.  Defaults to
            ``DEFAULT_MAP_NAME``.
        --pollster-identifier (str): Unique pollster identifier string.
            Defaults to ``DEFAULT_POLLSTER_IDENTIFIER``.
        --fieldwork-year-hint (int, optional): Explicit year to use when the
            fieldwork date string contains no year.
        --replace-rows (flag): Delete existing poll rows before inserting new
            ones.  Off by default (existing rows are preserved).
        --dry-run (flag): Print a preview of what would be imported without
            writing to the database.
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx-url", default=DEFAULT_XLSX_URL)
    parser.add_argument("--map-name", default=DEFAULT_MAP_NAME)
    parser.add_argument("--pollster-identifier", default=DEFAULT_POLLSTER_IDENTIFIER)
    parser.add_argument("--fieldwork-year-hint", type=int, default=None)
    parser.add_argument(
        "--replace-rows",
        action="store_true",
        help="Delete existing rows for the poll before inserting new ones",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    db = Database()
    print(f"Fetching XLSX: {args.xlsx_url}")

    plan = build_import_plan(
        db,
        xlsx_url=args.xlsx_url,
        map_name=args.map_name,
        pollster_identifier=args.pollster_identifier,
        fieldwork_year_hint=args.fieldwork_year_hint,
    )

    if args.dry_run:
        _cli_preview(plan)
        return

    result = commit_import_plan(db, plan, replace_rows=args.replace_rows)
    if result.created_pollster:
        print(f"created pollster: {args.pollster_identifier}")
    else:
        print(f"pollster exists: {args.pollster_identifier}")

    if result.created_poll:
        print(f"created poll: {result.poll_id}")
    else:
        print(f"poll exists: {result.poll_id}")

    if result.skipped_existing_rows:
        print(
            f"poll {result.poll_id} already has rows; "
            "use --replace-rows to overwrite"
        )
    else:
        if result.replaced_rows:
            print(f"deleted existing rows: {result.replaced_rows}")
        print(f"inserted poll rows: {result.inserted_rows}")


if __name__ == "__main__":
    main()
