#!/usr/bin/env python3
"""Import a Focaldata poll directly from an XLSX (or Google Sheet) URL."""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import select

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from db import Database
from models import Poll, PollRow, Pollster
from polls.import_types import PollImportResult

DEFAULT_XLSX_URL = "https://3959436.fs1.hubspotusercontent-na1.net/hubfs/3959436/Marketing%20content/Focaldata_January%20VI%20tables%2c%2016-19%20Jan%202026.xlsx"
DEFAULT_MAP_NAME = "UK Constituencies post 2022"
DEFAULT_POLLSTER_IDENTIFIER = "focaldata"
NATIONAL_KEY = "__national__"

MACRO_TO_INTERNAL_REGIONS = {
    "North of England": [
        "North East England",
        "North West England",
        "Yorkshire and The Humber",
    ],
    "Midlands": ["East Midlands", "West Midlands"],
    "South of England": ["East of England", "South East England", "South West England"],
    "Greater London": ["London"],
    "Wales": ["Wales"],
    "Scotland": ["Scotland"],
}

PARTY_NAME_MAP = {
    "Conservative": "Conservative",
    "Labour": "Labour",
    "Liberal Democrats": "Liberal Democrats",
    "Liberal Democrat": "Liberal Democrats",
    "Reform UK": "Reform UK",
    "Green": "Green",
    "Green Party": "Green",
    "Scottish National Party": "Scottish National Party",
    "Scottish National Party (SNP)": "Scottish National Party",
    "SNP": "Scottish National Party",
    "Plaid Cymru": "Plaid Cymru",
    "An independent candidate or other party (e.g. Workers Party / SDP / Yorkshire Party)": "Other",
    "An independent candidate or other party (e.g. Your Party / Workers Party / SDP)": "Other",
    "Another party": "Other",
    "Another party (e.g. Workers Party / SDP / Yorkshire Party)": "Other",
    "Other": "Other",
}


class ParsedPoll(BaseModel):
    """Parsed poll data extracted from source."""

    sample_size: int = Field(gt=0)
    fieldwork_start: date
    fieldwork_end: date
    party_macro_percentages: dict[str, dict[str, float]]


class PlannedPollRow(BaseModel):
    """A single poll row planned for DB insertion."""

    party_id: int
    party_name: str
    region_id: int | None
    region_name: str
    percentage: float = Field(ge=0, le=100)


class ImportPlan(BaseModel):
    """Full import plan: pollster, poll metadata, and rows."""

    pollster_identifier: str
    pollster_name: str
    pollster_id: int | None
    pollster_exists: bool
    regions_mapping: str
    map_id: int
    map_name: str
    source_url: str
    parsed: ParsedPoll
    poll_id: int | None
    poll_exists: bool
    rows: list[PlannedPollRow]


def _month_number(month_text: str) -> int | None:
    """Convert a month name or abbreviation to its integer number (1–12).

    Args:
        month_text: Month name or abbreviation, e.g. ``"Jan"``, ``"January"``,
            ``"Sept"``. Case-insensitive; leading/trailing whitespace and
            trailing periods are stripped.

    Returns:
        Integer month number (1–12), or ``None`` if the text is not recognised.
    """
    month_map = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }
    return month_map.get(month_text.strip().lower().rstrip("."))


def _infer_year(url: str, fallback: int | None = None) -> int:
    """Infer a four-digit poll year from a URL string.

    Searches for a path segment of the form ``/2025/`` first, then for any
    occurrence of a four-digit year starting with ``20``.

    Args:
        url: The source URL to search, e.g. an XLSX download link.
        fallback: Year to return when no year can be found in the URL.
            Optional; defaults to ``None``.

    Returns:
        The inferred year as an integer.

    Raises:
        ValueError: If no year can be found in the URL and no fallback is
            provided.
    """
    match = re.search(r"/(20\d{2})/", url)
    if match:
        return int(match.group(1))
    matches = re.findall(r"(20\d{2})", url)
    if matches:
        return int(matches[0])
    if fallback is not None:
        return fallback
    raise ValueError("Could not infer year from URL")


def _parse_fieldwork(fieldwork_text: str, default_year: int | None = None) -> tuple[date, date]:
    """Parse a human-readable fieldwork date range string into start/end dates.

    Handles four common formats (with or without an explicit year):

    * Same-month range with year:   ``"16-19 Jan 2026"``
    * Cross-month range with year:  ``"31 Jan - 3 Feb 2026"``
    * Same-month range, no year:    ``"16-19 Jan"``  (requires ``default_year``)
    * Cross-month range, no year:   ``"31 Jan - 3 Feb"`` (requires ``default_year``)

    Ordinal suffixes (``st``, ``nd``, ``rd``, ``th``) and em/en dashes are
    normalised automatically before matching.

    Args:
        fieldwork_text: Raw fieldwork date string as it appears in the source
            spreadsheet.
        default_year: Year to assume when the string contains no explicit year.
            Optional; defaults to ``None``.

    Returns:
        A ``(fieldwork_start, fieldwork_end)`` tuple of :class:`datetime.date`
        objects.

    Raises:
        ValueError: If the month names cannot be parsed or no pattern matches
            the normalised text.
    """
    normalized = re.sub(r"\s+", " ", fieldwork_text.strip().replace("–", "-").replace("—", "-"))
    normalized = re.sub(r"(\d)(st|nd|rd|th)", r"\1", normalized, flags=re.IGNORECASE)

    pattern_same_month = re.compile(r"(\d{1,2})\s*-\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})")
    match = pattern_same_month.search(normalized)
    if match:
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        year = int(match.group(4))
        if month is None:
            raise ValueError(f"Could not parse month in fieldwork: {fieldwork_text!r}")
        return date(year, month, day_start), date(year, month, day_end)

    pattern_cross_month = re.compile(
        r"(\d{1,2})\s+([A-Za-z]+)\s*-\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})"
    )
    match = pattern_cross_month.search(normalized)
    if match:
        day_start = int(match.group(1))
        month_start = _month_number(match.group(2))
        day_end = int(match.group(3))
        month_end = _month_number(match.group(4))
        year_end = int(match.group(5))
        if month_start is None or month_end is None:
            raise ValueError(f"Could not parse months in fieldwork: {fieldwork_text!r}")
        year_start = year_end - 1 if month_start > month_end else year_end
        return date(year_start, month_start, day_start), date(year_end, month_end, day_end)

    pattern_no_year = re.compile(r"(\d{1,2})\s*-\s*(\d{1,2})\s+([A-Za-z]+)")
    match = pattern_no_year.search(normalized)
    if match and default_year is not None:
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        if month is None:
            raise ValueError(f"Could not parse month in fieldwork: {fieldwork_text!r}")
        return date(default_year, month, day_start), date(default_year, month, day_end)

    pattern_cross_no_year = re.compile(r"(\d{1,2})\s+([A-Za-z]+)\s*-\s*(\d{1,2})\s+([A-Za-z]+)")
    match = pattern_cross_no_year.search(normalized)
    if match and default_year is not None:
        day_start = int(match.group(1))
        month_start = _month_number(match.group(2))
        day_end = int(match.group(3))
        month_end = _month_number(match.group(4))
        if month_start is None or month_end is None:
            raise ValueError(f"Could not parse months in fieldwork: {fieldwork_text!r}")
        year_start = default_year - 1 if month_start > month_end else default_year
        return date(year_start, month_start, day_start), date(default_year, month_end, day_end)

    raise ValueError(f"Could not parse fieldwork string: {fieldwork_text!r}")


def _google_sheet_export_url(url: str) -> str | None:
    """Convert a Google Sheets browser URL to a direct XLSX export URL.

    Extracts the spreadsheet ID from the path and the optional ``gid``
    query parameter (sheet tab identifier) to build an ``/export?format=xlsx``
    URL that can be fetched without a browser session.

    Args:
        url: A Google Sheets URL such as
            ``https://docs.google.com/spreadsheets/d/<ID>/edit#gid=<GID>``.

    Returns:
        An XLSX export URL string if ``url`` points to a Google Sheet, or
        ``None`` if it does not match the expected domain/path pattern.
    """
    parsed = urlparse(url)
    if "docs.google.com" not in parsed.netloc or "/spreadsheets/d/" not in parsed.path:
        return None
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", parsed.path)
    if not match:
        return None
    spreadsheet_id = match.group(1)
    query = parse_qs(parsed.query)
    gid = query.get("gid", [None])[0]
    export = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    if gid:
        export += f"&gid={gid}"
    return export


def extract_workbook(xlsx_url: str) -> Any:
    """Download an XLSX file and return an openpyxl workbook.

    If ``xlsx_url`` looks like a Google Sheets browser URL the function first
    tries the derived XLSX export URL, then falls back to the original URL.
    Only a response whose body starts with the ZIP/PK magic bytes is accepted
    as a valid XLSX payload.

    Args:
        xlsx_url: Direct XLSX download URL or a Google Sheets browser URL.

    Returns:
        An :class:`openpyxl.Workbook` loaded in ``data_only`` mode (formulas
        resolved to their cached values).

    Raises:
        ValueError: If no candidate URL returns a valid XLSX payload, with a
            summary of all per-URL errors.
    """
    candidate_urls = [xlsx_url]
    gs_export = _google_sheet_export_url(xlsx_url)
    if gs_export:
        candidate_urls.insert(0, gs_export)

    payload = None
    errors: list[str] = []
    for candidate in dict.fromkeys(candidate_urls):
        req = Request(candidate, headers={"User-Agent": "Mozilla/5.0 (compatible; poll-importer/1.0)"})
        try:
            with urlopen(req, timeout=50) as response:
                data = response.read()
            if data.startswith(b"PK"):
                payload = data
                break
            errors.append(f"non-xlsx payload at {candidate}")
        except Exception as exc:
            errors.append(f"{candidate}: {exc}")

    if payload is None:
        raise ValueError("Could not fetch XLSX payload: " + " | ".join(errors))

    return load_workbook(filename=BytesIO(payload), data_only=True)


def _cell_text(value: object) -> str:
    """Return a cell value as a stripped string, or an empty string for ``None``.

    Args:
        value: Raw cell value from openpyxl (may be ``None``, a number, a
            string, or another scalar type).

    Returns:
        String representation of ``value`` with leading/trailing whitespace
        removed, or ``""`` if ``value`` is ``None``.
    """
    if value is None:
        return ""
    return str(value).strip()


def _to_percentage(value: object) -> float:
    """Convert a cell value to a whole-number percentage (0–100 scale).

    Values already in the 0–1 range (i.e. Excel-stored proportions) are
    multiplied by 100 before rounding to the nearest integer.

    Args:
        value: Raw numeric cell value. Must not be ``None``.

    Returns:
        Percentage as a float rounded to the nearest integer (e.g. ``42.0``).

    Raises:
        ValueError: If ``value`` is ``None``.
    """
    if value is None:
        raise ValueError("Encountered empty percentage cell")
    number = float(value)  # type: ignore[arg-type]
    pct = number * 100.0 if 0.0 <= number <= 1.0 else number
    return float(int(round(pct)))


def _to_percentage_or_zero(value: object) -> float:
    """Convert a cell value to a percentage, returning ``0.0`` for empty/dash cells.

    Treats ``None``, empty strings, and common dash characters (``-``, ``–``,
    ``—``) as zero rather than raising an error. All other values are delegated
    to :func:`_to_percentage`.

    Args:
        value: Raw cell value from openpyxl (may be ``None``, a number, or a
            string placeholder such as ``"-"``).

    Returns:
        Percentage as a float rounded to the nearest integer, or ``0.0`` for
        blank/placeholder cells.
    """
    if value is None:
        return 0.0
    if isinstance(value, str):
        cleaned = value.strip()
        if cleaned in {"", "-", "–", "—"}:
            return 0.0
    return _to_percentage(value)


def _canonical_party(label: str) -> str | None:
    """Resolve a raw party label from the spreadsheet to a canonical party name.

    First performs a direct lookup in :data:`PARTY_NAME_MAP`. If not found,
    applies case-insensitive substring heuristics (e.g. ``"reform"`` →
    ``"Reform UK"``).

    Args:
        label: Raw party label string as it appears in a spreadsheet cell.

    Returns:
        A canonical party name string (one of the values in
        :data:`PARTY_NAME_MAP`), or ``None`` if the label cannot be resolved.
    """
    direct = PARTY_NAME_MAP.get(label)
    if direct is not None:
        return direct
    lowered = label.lower()
    if "reform" in lowered:
        return "Reform UK"
    if "liberal" in lowered and "dem" in lowered:
        return "Liberal Democrats"
    if "green" in lowered:
        return "Green"
    if "conserv" in lowered:
        return "Conservative"
    if lowered.startswith("labour"):
        return "Labour"
    if "scottish national" in lowered or lowered == "snp":
        return "Scottish National Party"
    if "plaid" in lowered:
        return "Plaid Cymru"
    if "independent" in lowered or "another party" in lowered or lowered == "other":
        return "Other"
    return None


def _parse_info_sheet(workbook: Any, default_year: int, fieldwork_label: str | None = None) -> tuple[date, date, int]:
    """Extract fieldwork dates and sample size from the workbook's info sheet.

    Looks for a sheet named ``"Info"`` first, falling back to the first sheet.
    Searches the first 80 rows for ``"dates conducted"`` and ``"sample size"``
    labels in columns B/C (and A/B as a fallback). If sample size is still not
    found, falls back to scanning the tables sheet for an ``"Unweighted sample"``
    row. If fieldwork dates are still missing and ``fieldwork_label`` is
    provided, that string is used directly.

    Args:
        workbook: An openpyxl :class:`~openpyxl.Workbook` object loaded in
            ``data_only`` mode.
        default_year: Four-digit year used when the fieldwork string contains
            no explicit year.
        fieldwork_label: Optional override for the fieldwork date string when
            it cannot be found in the sheet. Optional; defaults to ``None``.

    Returns:
        A ``(fieldwork_start, fieldwork_end, sample_size)`` tuple.

    Raises:
        ValueError: If fieldwork dates or sample size cannot be found in the
            workbook.
    """
    sheet = workbook["Info"] if "Info" in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    fieldwork_raw = None
    sample_size = None

    for row in range(1, 80):
        label = _cell_text(sheet.cell(row, 2).value).lower()
        value = sheet.cell(row, 3).value
        if "dates conducted" in label:
            fieldwork_raw = _cell_text(value)
        if "sample size" in label and value is not None:
            if isinstance(value, (int, float)):
                sample_size = int(round(float(value)))
            else:
                digits = re.sub(r"[^0-9]", "", _cell_text(value))
                if digits:
                    sample_size = int(digits)

    if fieldwork_raw is None:
        for row in range(1, 80):
            label = _cell_text(sheet.cell(row, 1).value).lower()
            if "dates conducted" in label:
                fieldwork_raw = _cell_text(sheet.cell(row, 2).value)
                if not fieldwork_raw:
                    fieldwork_raw = _cell_text(sheet.cell(row, 3).value)

    if sample_size is None:
        for row in range(1, 80):
            label = _cell_text(sheet.cell(row, 1).value).lower()
            if "sample size" in label:
                for col in range(2, 6):
                    value = sheet.cell(row, col).value
                    if value is None:
                        continue
                    if isinstance(value, (int, float)):
                        sample_size = int(round(float(value)))
                        break
                    digits = re.sub(r"[^0-9]", "", _cell_text(value))
                    if digits:
                        sample_size = int(digits)
                        break
                if sample_size is not None:
                    break

    if sample_size is None:
        table_sheet = _find_tables_sheet(workbook)
        for row in range(1, min(table_sheet.max_row, 40) + 1):
            first = _cell_text(table_sheet.cell(row, 1).value).lower()
            if first.startswith("unweighted sample"):
                for col in range(2, 8):
                    val = table_sheet.cell(row, col).value
                    if isinstance(val, (int, float)):
                        sample_size = int(round(float(val)))
                        break
                if sample_size is not None:
                    break

    if not fieldwork_raw:
        if fieldwork_label:
            fieldwork_raw = fieldwork_label
        else:
            raise ValueError("Dates conducted not found in workbook")
    if sample_size is None:
        raise ValueError("Sample size not found in workbook")

    fieldwork_start, fieldwork_end = _parse_fieldwork(fieldwork_raw, default_year=default_year)
    return fieldwork_start, fieldwork_end, sample_size


def _find_tables_sheet(workbook: Any) -> Any:
    """Return the cross-tabulation sheet from the workbook.

    Identifies the sheet by matching its name (case-insensitive) against
    ``"tables"``, any name containing ``"table"`` or ``"voting intention"``,
    or ``"respondents"``.

    Args:
        workbook: An openpyxl :class:`~openpyxl.Workbook` object.

    Returns:
        The matching openpyxl worksheet object.

    Raises:
        ValueError: If no sheet name matches the expected patterns.
    """
    for name in workbook.sheetnames:
        lowered = name.lower()
        if lowered == "tables" or "table" in lowered or "voting intention" in lowered or "respondents" in lowered:
            return workbook[name]
    raise ValueError("Could not find tables sheet")


def _find_vi_section_start(sheet: Any) -> int:
    """Locate the first row of the voting intention section in a tables sheet.

    Scans the first 500 rows of column A looking for one of three known
    question-header phrasings (tried in priority order):

    1. Contains ``"combined voting intention"`` and ``"excluding"``.
    2. Contains ``"general election held in the next few weeks"``.
    3. Contains both ``"general election"`` and ``"vote for"``.

    Args:
        sheet: An openpyxl worksheet object for the cross-tabulation sheet.

    Returns:
        The 1-based row number of the matched header row.

    Raises:
        ValueError: If none of the expected header phrasings are found within
            the first 500 rows.
    """
    for row in range(1, min(sheet.max_row, 500) + 1):
        value = _cell_text(sheet.cell(row, 1).value).lower()
        if "combined voting intention" in value and "excluding" in value:
            return row

    for row in range(1, min(sheet.max_row, 500) + 1):
        value = _cell_text(sheet.cell(row, 1).value).lower()
        if "general election held in the next few weeks" in value:
            return row

    for row in range(1, min(sheet.max_row, 500) + 1):
        value = _cell_text(sheet.cell(row, 1).value).lower()
        if "general election" in value and "vote for" in value:
            return row

    raise ValueError("Could not find voting intention section in tables sheet")


def _parse_party_macro_percentages(workbook: Any) -> dict[str, dict[str, float]]:
    """Extract national and macro-region vote-share percentages for each party.

    Locates the voting intention section via :func:`_find_vi_section_start`,
    then scans upward for a region header row (columns 11–24) that contains
    known macro-region names from :data:`MACRO_TO_INTERNAL_REGIONS`. For each
    data row, resolves the party label via :func:`_canonical_party` and reads
    the national total plus one percentage per macro-region column.

    Optional parties (``"Scottish National Party"``, ``"Plaid Cymru"``,
    ``"Other"``) default to zero if absent from the sheet.

    Args:
        workbook: An openpyxl :class:`~openpyxl.Workbook` object loaded in
            ``data_only`` mode.

    Returns:
        A nested dict mapping canonical party name → region key → percentage,
        where the special key :data:`NATIONAL_KEY` (``"__national__"``) holds
        the national figure and all other keys are macro-region names from
        :data:`MACRO_TO_INTERNAL_REGIONS`.

    Raises:
        ValueError: If any of the required party rows
            (Conservative, Labour, Liberal Democrats, Reform UK, Green,
            Scottish National Party, Plaid Cymru, Other) are missing from the
            parsed data.
    """
    sheet = _find_tables_sheet(workbook)
    start_row = _find_vi_section_start(sheet)

    region_header_row = None
    for row in range(start_row + 1, min(sheet.max_row, start_row + 12)):
        headers = {_cell_text(sheet.cell(row, col).value) for col in range(11, 25)}
        if any(header in MACRO_TO_INTERNAL_REGIONS for header in headers):
            region_header_row = row
            break

    macro_columns: dict[str, int] = {}
    if region_header_row is not None:
        for col in range(11, 30):
            header = _cell_text(sheet.cell(region_header_row, col).value)
            if header in MACRO_TO_INTERNAL_REGIONS:
                macro_columns[header] = col

    parsed: dict[str, dict[str, float]] = {}
    for row in range(start_row, min(sheet.max_row, start_row + 180)):
        label_col1 = _cell_text(sheet.cell(row, 1).value)
        label_col2 = _cell_text(sheet.cell(row, 2).value)
        if not label_col1 and not label_col2:
            continue

        lowered = label_col1.lower()
        if lowered.startswith("column n") or lowered.startswith("column population"):
            break
        if lowered.startswith("q") and row > start_row + 3:
            break

        canonical = None
        value = None

        canonical_from_col1 = _canonical_party(label_col1)
        if canonical_from_col1 is not None:
            canonical = canonical_from_col1
            candidate = sheet.cell(row, 2).value
            if isinstance(candidate, (int, float)):
                value = candidate
            else:
                candidate = sheet.cell(row, 3).value
                if isinstance(candidate, (int, float)):
                    value = candidate

        canonical_from_col2 = _canonical_party(label_col2)
        if canonical is None and canonical_from_col2 is not None:
            canonical = canonical_from_col2
            candidate = sheet.cell(row, 3).value
            if isinstance(candidate, (int, float)):
                value = candidate
            else:
                candidate = sheet.cell(row, 2).value
                if isinstance(candidate, (int, float)):
                    value = candidate

        if canonical is None or value is None:
            continue

        macro_values = {NATIONAL_KEY: _to_percentage(value)}
        for macro_name, col in macro_columns.items():
            macro_values[macro_name] = _to_percentage_or_zero(sheet.cell(row, col).value)
        parsed[canonical] = macro_values

    for optional_party in ["Scottish National Party", "Plaid Cymru", "Other"]:
        parsed.setdefault(optional_party, {NATIONAL_KEY: 0.0})
        for macro_name in MACRO_TO_INTERNAL_REGIONS:
            parsed[optional_party].setdefault(macro_name, 0.0)

    required = {
        "Conservative",
        "Labour",
        "Liberal Democrats",
        "Reform UK",
        "Green",
        "Scottish National Party",
        "Plaid Cymru",
        "Other",
    }
    missing = sorted(required - set(parsed.keys()))
    if missing:
        raise ValueError(f"Missing expected party rows in workbook: {missing}")

    return parsed


def parse_poll(
    workbook: Any,
    *,
    source_url: str,
    year_hint: int | None = None,
    fieldwork_label: str | None = None,
) -> ParsedPoll:
    """Parse a Focaldata workbook into a :class:`ParsedPoll` data object.

    Combines metadata extraction (fieldwork dates, sample size) from the info
    sheet and vote-share extraction from the tables sheet.

    Args:
        workbook: An openpyxl :class:`~openpyxl.Workbook` object loaded in
            ``data_only`` mode.
        source_url: URL the workbook was downloaded from; used to infer the
            poll year when no explicit year appears in the fieldwork string.
        year_hint: Optional four-digit year fallback used when the year cannot
            be inferred from ``source_url``. Defaults to ``None``.
        fieldwork_label: Optional raw fieldwork date string to use when the
            info sheet does not contain one. Defaults to ``None``.

    Returns:
        A :class:`ParsedPoll` instance populated with sample size, fieldwork
        dates, and per-party macro-region percentages.

    Raises:
        ValueError: Propagated from :func:`_infer_year`,
            :func:`_parse_info_sheet`, or :func:`_parse_party_macro_percentages`
            if required data cannot be extracted.
    """
    year = _infer_year(source_url, fallback=year_hint)
    fieldwork_start, fieldwork_end, sample_size = _parse_info_sheet(workbook, year, fieldwork_label=fieldwork_label)
    party_macro_percentages = _parse_party_macro_percentages(workbook)
    return ParsedPoll(
        sample_size=sample_size,
        fieldwork_start=fieldwork_start,
        fieldwork_end=fieldwork_end,
        party_macro_percentages=party_macro_percentages,
    )


def _find_existing_poll(db: Database, pollster_id: int, map_id: int, parsed: ParsedPoll) -> Poll | None:
    """Look up a poll in the database matching the given metadata.

    Matches on pollster ID, map ID, fieldwork start/end dates, and sample size.

    Args:
        db: Active :class:`~db.Database` connection.
        pollster_id: Primary key of the pollster record.
        map_id: Primary key of the constituency map record.
        parsed: Parsed poll data supplying fieldwork dates and sample size to
            match against.

    Returns:
        The matching :class:`~models.Poll` ORM instance, or ``None`` if no
        matching record exists.
    """
    with db.session() as session:
        return session.execute(
            select(Poll).where(
                Poll.pollster_id == pollster_id,
                Poll.map_id == map_id,
                Poll.fieldwork_start == parsed.fieldwork_start,
                Poll.fieldwork_end == parsed.fieldwork_end,
                Poll.sample_size == parsed.sample_size,
            )
        ).scalar_one_or_none()


def build_import_plan(
    db: Database,
    *,
    xlsx_url: str = DEFAULT_XLSX_URL,
    map_name: str = DEFAULT_MAP_NAME,
    pollster_identifier: str = DEFAULT_POLLSTER_IDENTIFIER,
    year_hint: int | None = None,
    fieldwork_label: str | None = None,
) -> ImportPlan:
    """Build a full import plan from a Focaldata XLSX without writing to the DB.

    Downloads and parses the workbook, resolves pollster and map records,
    expands macro-region percentages to individual DB region rows, and checks
    whether the poll already exists. The returned plan can be previewed or
    passed directly to :func:`commit_import_plan`.

    Args:
        db: Active :class:`~db.Database` connection used for lookups.
        xlsx_url: URL of the XLSX file (or Google Sheets browser URL) to
            import. Defaults to :data:`DEFAULT_XLSX_URL`.
        map_name: Name of the constituency map to associate the poll with.
            Must exist in the database. Defaults to :data:`DEFAULT_MAP_NAME`.
        pollster_identifier: Short identifier string for the pollster record
            (e.g. ``"focaldata"``). Defaults to
            :data:`DEFAULT_POLLSTER_IDENTIFIER`.
        year_hint: Optional four-digit year used when the year cannot be
            inferred from ``xlsx_url``. Defaults to ``None``.
        fieldwork_label: Optional raw fieldwork date string passed through to
            the parser when the info sheet lacks one. Defaults to ``None``.

    Returns:
        An :class:`ImportPlan` describing what would be created or updated,
        including all :class:`PlannedPollRow` objects.

    Raises:
        ValueError: If the map is not found, any required party is missing from
            the database, or the workbook cannot be parsed.
    """
    poll_map = db.get_map_by_name(map_name)
    if poll_map is None:
        raise ValueError(f"Map not found: {map_name!r}")

    workbook = extract_workbook(xlsx_url)
    parsed = parse_poll(
        workbook,
        source_url=xlsx_url,
        year_hint=year_hint,
        fieldwork_label=fieldwork_label,
    )

    pollster = db.get_pollster_by_identifier(pollster_identifier)
    pollster_exists = pollster is not None

    party_by_name = {party.name: party for party in db.get_all_parties()}
    missing_parties = [
        party_name
        for party_name in sorted(set(PARTY_NAME_MAP.values()))
        if party_name not in party_by_name
    ]
    if missing_parties:
        raise ValueError(
            "Missing parties in database (run party importer first): "
            f"{missing_parties}"
        )

    regions = db.get_regions_for_map(poll_map.id)
    region_ids_by_name: dict[str, int] = {region.name: region.id for region in regions}

    macro_to_region_ids: dict[str, list[int]] = {}
    for macro_name, internal_names in MACRO_TO_INTERNAL_REGIONS.items():
        ids: list[int] = []
        for internal_name in internal_names:
            region_id = region_ids_by_name.get(internal_name)
            if region_id is not None:
                ids.append(region_id)
        macro_to_region_ids[macro_name] = ids

    rows: list[PlannedPollRow] = []
    for party_name, macro_values in parsed.party_macro_percentages.items():
        rows.append(
            PlannedPollRow(
                party_id=party_by_name[party_name].id,
                party_name=party_name,
                region_id=None,
                region_name="National",
                percentage=macro_values.get(NATIONAL_KEY, 0.0),
            )
        )

        percentages_by_region_id: dict[int, float] = {}
        for macro_name, region_ids in macro_to_region_ids.items():
            macro_pct = macro_values.get(macro_name, 0.0)
            for region_id in region_ids:
                percentages_by_region_id[region_id] = macro_pct

        for region in regions:
            rows.append(
                PlannedPollRow(
                    party_id=party_by_name[party_name].id,
                    party_name=party_name,
                    region_id=region.id,
                    region_name=region.name,
                    percentage=percentages_by_region_id.get(region.id, 0.0),
                )
            )

    existing_poll = (
        _find_existing_poll(db, pollster.id, poll_map.id, parsed)
        if pollster is not None
        else None
    )

    return ImportPlan(
        pollster_identifier=pollster_identifier,
        pollster_name=(pollster.name if pollster else "Focaldata"),
        pollster_id=(pollster.id if pollster else None),
        pollster_exists=pollster_exists,
        regions_mapping="",
        map_id=poll_map.id,
        map_name=poll_map.name,
        source_url=xlsx_url,
        parsed=parsed,
        poll_id=(existing_poll.id if existing_poll else None),
        poll_exists=existing_poll is not None,
        rows=rows,
    )


def commit_import_plan(
    db: Database,
    plan: ImportPlan,
    *,
    replace_rows: bool = False,
) -> PollImportResult:
    """Persist an :class:`ImportPlan` to the database.

    Creates the pollster and/or poll records if they do not already exist.
    If the poll already has rows and ``replace_rows`` is ``False``, the
    function returns early without inserting any rows. When ``replace_rows``
    is ``True``, all existing rows for the poll are deleted before the new
    rows from the plan are inserted.

    Side effects:
        - May insert a :class:`~models.Pollster` row.
        - May insert a :class:`~models.Poll` row (or update its ``source_url``).
        - May delete and re-insert :class:`~models.PollRow` rows.

    Args:
        db: Active :class:`~db.Database` connection.
        plan: The :class:`ImportPlan` produced by :func:`build_import_plan`.
        replace_rows: If ``True``, delete existing poll rows before inserting
            new ones. Defaults to ``False``.

    Returns:
        A :class:`~polls.import_types.PollImportResult` summarising what was
        created, replaced, or skipped.

    Raises:
        ValueError: If the pollster lookup fails during commit (should not
            happen in normal usage).
    """
    if plan.pollster_exists:
        pollster = db.get_pollster_by_identifier(plan.pollster_identifier)
        if pollster is None:
            raise ValueError("Pollster lookup failed during commit")
        pollster_id = pollster.id
        created_pollster = False
    else:
        created = db.add_pollster(
            name=plan.pollster_name,
            identifier=plan.pollster_identifier,
            weight=1.0,
            regions_mapping=plan.regions_mapping,
        )
        pollster_id = created.id
        created_pollster = True

    existing_poll = _find_existing_poll(db, pollster_id, plan.map_id, plan.parsed)
    if existing_poll is None:
        created_poll = db.add_poll(
            pollster_id=pollster_id,
            map_id=plan.map_id,
            fieldwork_start=plan.parsed.fieldwork_start,
            fieldwork_end=plan.parsed.fieldwork_end,
            sample_size=plan.parsed.sample_size,
            source_url=plan.source_url,
        )
        poll_id = created_poll.id
        created_poll_row = True
    else:
        with db.session() as session:
            db_poll = session.get(Poll, existing_poll.id)
            if db_poll is not None and db_poll.source_url != plan.source_url:
                db_poll.source_url = plan.source_url
        poll_id = existing_poll.id
        created_poll_row = False

    with db.session() as session:
        existing_rows = session.execute(
            select(PollRow).where(PollRow.poll_id == poll_id)
        ).scalars().all()

        if existing_rows and not replace_rows:
            return PollImportResult(
                created_pollster=created_pollster,
                created_poll=created_poll_row,
                poll_id=poll_id,
                inserted_rows=0,
                replaced_rows=0,
                skipped_existing_rows=True,
            )

        replaced_rows = 0
        if existing_rows and replace_rows:
            for existing_row in existing_rows:
                session.delete(existing_row)
                replaced_rows += 1

        for row in plan.rows:
            session.add(
                PollRow(
                    poll_id=poll_id,
                    region_id=row.region_id,
                    party_id=row.party_id,
                    percentage=row.percentage,
                )
            )

    return PollImportResult(
        created_pollster=created_pollster,
        created_poll=created_poll_row,
        poll_id=poll_id,
        inserted_rows=len(plan.rows),
        replaced_rows=replaced_rows,
        skipped_existing_rows=False,
    )


def _cli_preview(plan: ImportPlan) -> None:
    """Print a human-readable dry-run summary of an :class:`ImportPlan` to stdout.

    Outputs the parsed fieldwork dates and sample size, whether the pollster
    and poll already exist, and every planned poll row with its party, region,
    and percentage.

    Args:
        plan: The :class:`ImportPlan` to preview.
    """
    print(
        "Parsed poll: "
        f"fieldwork={plan.parsed.fieldwork_start} to {plan.parsed.fieldwork_end}, "
        f"sample={plan.parsed.sample_size}"
    )
    if plan.pollster_exists:
        print(f"pollster exists: {plan.pollster_identifier}")
    else:
        print(f"[dry-run] would create pollster: {plan.pollster_identifier}")

    if plan.poll_exists and plan.poll_id is not None:
        print(f"poll exists: {plan.poll_id}")
    else:
        print("[dry-run] would create poll")

    for row in plan.rows:
        print(
            "[dry-run] would insert row: "
            f"party={row.party_name}, region={row.region_name}, "
            f"region_id={row.region_id}, pct={row.percentage:.2f}"
        )


def main() -> None:
    """CLI entry point for importing a Focaldata poll from an XLSX URL.

    Parses command-line arguments, builds an :class:`ImportPlan`, and either
    prints a dry-run preview or commits the plan to the database.

    CLI arguments:
        --xlsx-url (str): URL of the XLSX file to import. Defaults to
            :data:`DEFAULT_XLSX_URL`.
        --map-name (str): Name of the constituency map. Defaults to
            :data:`DEFAULT_MAP_NAME`.
        --pollster-identifier (str): Short pollster identifier string. Defaults
            to :data:`DEFAULT_POLLSTER_IDENTIFIER`.
        --year-hint (int): Optional four-digit year hint for date parsing.
        --replace-rows (flag): Delete existing poll rows before inserting new
            ones.
        --dry-run (flag): Print the import plan without writing to the database.
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx-url", default=DEFAULT_XLSX_URL)
    parser.add_argument("--map-name", default=DEFAULT_MAP_NAME)
    parser.add_argument("--pollster-identifier", default=DEFAULT_POLLSTER_IDENTIFIER)
    parser.add_argument("--year-hint", type=int, default=None)
    parser.add_argument(
        "--replace-rows",
        action="store_true",
        help="Delete existing rows for the poll before inserting new ones",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    db = Database()
    print(f"Fetching XLSX: {args.xlsx_url}")

    plan = build_import_plan(
        db,
        xlsx_url=args.xlsx_url,
        map_name=args.map_name,
        pollster_identifier=args.pollster_identifier,
        year_hint=args.year_hint,
    )

    if args.dry_run:
        _cli_preview(plan)
        return

    result = commit_import_plan(db, plan, replace_rows=args.replace_rows)
    if result.created_pollster:
        print(f"created pollster: {args.pollster_identifier}")
    else:
        print(f"pollster exists: {args.pollster_identifier}")

    if result.created_poll:
        print(f"created poll: {result.poll_id}")
    else:
        print(f"poll exists: {result.poll_id}")

    if result.skipped_existing_rows:
        print(
            f"poll {result.poll_id} already has rows; "
            "use --replace-rows to overwrite"
        )
    else:
        if result.replaced_rows:
            print(f"deleted existing rows: {result.replaced_rows}")
        print(f"inserted poll rows: {result.inserted_rows}")


if __name__ == "__main__":
    main()
