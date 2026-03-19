#!/usr/bin/env python3
"""Import a More in Common poll directly from an XLSX URL."""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.request import urlopen

from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import select

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from db import Database
from models import Poll, PollRow, Pollster
from polls.import_types import PollImportResult

DEFAULT_XLSX_URL = "https://www.moreincommon.org.uk/media/dshdwjt0/voting-intention-and-trackers-10-feb.xlsx"
DEFAULT_MAP_NAME = "UK Constituencies post 2022"
DEFAULT_POLLSTER_IDENTIFIER = "more_in_common"

NATIONAL_KEY = "__national__"

PARTY_NAME_MAP = {
    "Conservatives": "Conservative",
    "Conservative": "Conservative",
    "Labour": "Labour",
    "Liberal Democrat": "Liberal Democrats",
    "Liberal Democrats": "Liberal Democrats",
    "Reform UK": "Reform UK",
    "The Green Party": "Green",
    "Green Party": "Green",
    "Green": "Green",
    "Scottish National Party": "Scottish National Party",
    "Scottish National Party (SNP)": "Scottish National Party",
    "The SNP": "Scottish National Party",
    "SNP": "Scottish National Party",
    "Plaid Cymru": "Plaid Cymru",
    "Another Party/ Independent": "Other",
    "Another party/ Independent": "Other",
    "Another Party/Independent": "Other",
    "Another party/Independent": "Other",
    "Another party/Independent candidate": "Other",
    "Another party/Independent Candidate": "Other",
    "Another party": "Other",
    "Other": "Other",
}

REGION_HEADER_TO_INTERNAL = {
    "East Midlands": "East Midlands",
    "East of England": "East of England",
    "Greater London": "London",
    "London": "London",
    "North East England": "North East England",
    "North West England": "North West England",
    "Scotland": "Scotland",
    "South East England": "South East England",
    "South West England": "South West England",
    "Wales": "Wales",
    "West Midlands": "West Midlands",
    "Yorkshire and the Humber": "Yorkshire and The Humber",
}


class ParsedPoll(BaseModel):
    """Parsed poll data extracted from source."""

    sample_size: int = Field(gt=0)
    fieldwork_start: date
    fieldwork_end: date
    party_region_percentages: dict[str, dict[str, float]]


class PlannedPollRow(BaseModel):
    """A single poll row planned for DB insertion."""

    party_id: int
    party_name: str
    region_id: int | None
    region_name: str
    percentage: float = Field(ge=0, le=100)


class ImportPlan(BaseModel):
    """Full import plan: pollster, poll metadata, and rows."""

    pollster_identifier: str
    pollster_name: str
    pollster_id: int | None
    pollster_exists: bool
    regions_mapping: str
    map_id: int
    map_name: str
    source_url: str
    parsed: ParsedPoll
    poll_id: int | None
    poll_exists: bool
    rows: list[PlannedPollRow]


def normalize_name(value: str) -> str:
    """Collapse internal whitespace and lowercase a string for comparison."""
    return " ".join(value.strip().split()).lower()


def _month_number(month_text: str) -> int | None:
    """Convert a month name or abbreviation to its 1-based integer index.

    Args:
        month_text: Month name or abbreviation, e.g. ``"Jan"``, ``"january"``,
            ``"Feb."``.  Case-insensitive; trailing periods are ignored.

    Returns:
        Integer month number (1–12), or ``None`` if the string is not
        recognised.
    """
    cleaned = month_text.strip().lower().rstrip(".")
    month_map = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }
    return month_map.get(cleaned)


def parse_fieldwork(value: str, *, default_year: int | None = None) -> tuple[date, date]:
    """Parse a human-readable fieldwork date string into a start/end date pair.

    Handles several common formats used in More in Common XLSX files:

    - Cross-month ranges: ``"28 Jan - 3 Feb 2025"``
    - Same-month ranges with year: ``"3-7 February 2025"``
    - Same-month ranges without year: ``"3-7 February"`` (requires ``default_year``)
    - Single days with year: ``"5 March 2025"``
    - Single days without year: ``"5 March"`` (requires ``default_year``)

    En-dashes and em-dashes are normalised to hyphens before matching.

    Args:
        value: Raw fieldwork date string extracted from the workbook.
        default_year: Year to use when the string contains no four-digit year.
            Optional; if ``None`` and the string has no year, parsing fails.

    Returns:
        Tuple of ``(fieldwork_start, fieldwork_end)`` as :class:`datetime.date`
        objects.  For a single-day poll both elements are the same date.

    Raises:
        ValueError: If the string does not match any recognised pattern.
    """
    normalized = re.sub(r"\s+", " ", value.strip().replace("–", "-").replace("—", "-"))

    cross_month_pattern = re.compile(
        r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z.]+)\s*-\s*(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z.]+)\s+(\d{4})"
    )
    range_same_month_pattern = re.compile(
        r"(\d{1,2})(?:st|nd|rd|th)?\s*-\s*(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z.]+)\s+(\d{4})"
    )
    range_same_month_no_year_pattern = re.compile(
        r"(\d{1,2})(?:st|nd|rd|th)?\s*-\s*(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z.]+)"
    )
    single_day_pattern = re.compile(r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z.]+)\s+(\d{4})")
    single_day_no_year_pattern = re.compile(r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z.]+)$")

    match = cross_month_pattern.search(normalized)
    if match:
        day_start = int(match.group(1))
        month_start = _month_number(match.group(2))
        day_end = int(match.group(3))
        month_end = _month_number(match.group(4))
        year_end = int(match.group(5))
        if month_start is None or month_end is None:
            raise ValueError(f"Could not parse fieldwork string: {value!r}")
        year_start = year_end - 1 if month_start > month_end else year_end
        return date(year_start, month_start, day_start), date(year_end, month_end, day_end)

    match = range_same_month_pattern.search(normalized)
    if match:
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        year = int(match.group(4))
        if month is None:
            raise ValueError(f"Could not parse fieldwork string: {value!r}")
        return date(year, month, day_start), date(year, month, day_end)

    match = range_same_month_no_year_pattern.search(normalized)
    if match and default_year is not None:
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        if month is None:
            raise ValueError(f"Could not parse fieldwork string: {value!r}")
        return date(default_year, month, day_start), date(default_year, month, day_end)

    match = single_day_pattern.search(normalized)
    if match:
        day = int(match.group(1))
        month = _month_number(match.group(2))
        year = int(match.group(3))
        if month is None:
            raise ValueError(f"Could not parse fieldwork string: {value!r}")
        parsed_day = date(year, month, day)
        return parsed_day, parsed_day

    match = single_day_no_year_pattern.search(normalized)
    if match and default_year is not None:
        day = int(match.group(1))
        month = _month_number(match.group(2))
        if month is None:
            raise ValueError(f"Could not parse fieldwork string: {value!r}")
        parsed_day = date(default_year, month, day)
        return parsed_day, parsed_day

    raise ValueError(f"Could not parse fieldwork string: {value!r}")


def extract_workbook(xlsx_url: str) -> Any:
    """Download an XLSX file from a URL and return the parsed openpyxl workbook.

    Args:
        xlsx_url: Fully-qualified URL pointing to the XLSX file.

    Returns:
        An openpyxl ``Workbook`` object with ``data_only=True`` (formula
        results are read instead of formula strings).
    """
    payload = urlopen(xlsx_url).read()
    return load_workbook(filename=BytesIO(payload), data_only=True)


def _cell_text(value: object) -> str:
    """Return the string representation of a cell value, stripped of whitespace.

    Args:
        value: Raw cell value from openpyxl (may be ``None``, ``int``,
            ``float``, ``str``, or other types).

    Returns:
        Stripped string, or an empty string if ``value`` is ``None``.
    """
    if value is None:
        return ""
    return str(value).strip()


def _find_label_value(ws: Any, label_fragment: str) -> str | None:
    """Search a worksheet for a label and return the first non-empty value to its right.

    Scans rows 1–44 and columns 1–7 for a cell whose text contains
    ``label_fragment`` (case-insensitive).  Once found, checks up to three
    cells to the right and returns the first non-empty one.

    Args:
        ws: An openpyxl ``Worksheet`` object.
        label_fragment: Substring to search for in cell text, e.g.
            ``"Fieldwork"`` or ``"Sample size"``.

    Returns:
        The text of the first non-empty cell to the right of the label, or
        ``None`` if no match is found.
    """
    needle = label_fragment.lower()
    for row in range(1, 45):
        for col in range(1, 8):
            label = _cell_text(ws.cell(row, col).value)
            if not label:
                continue
            if needle in label.lower():
                for offset in range(1, 4):
                    candidate = _cell_text(ws.cell(row, col + offset).value)
                    if candidate:
                        return candidate
    return None


def _as_int(value: str) -> int:
    """Extract a non-negative integer from a string by stripping non-digit characters.

    Args:
        value: Raw string that may contain formatting characters such as commas
            or spaces (e.g. ``"1,234"`` or ``"n=2000"``).

    Returns:
        Integer parsed from the digit characters in ``value``.

    Raises:
        ValueError: If ``value`` contains no digit characters.
    """
    digits = re.sub(r"[^0-9]", "", value)
    if not digits:
        raise ValueError(f"Could not parse sample size from value: {value!r}")
    return int(digits)


def _to_percentage(value: object) -> float:
    """Convert a cell value to a rounded percentage (0–100 scale).

    Values already in the 0–100 range are used as-is; values in the 0–1
    range (as stored by Excel for proportions) are multiplied by 100.  The
    result is rounded to the nearest whole number.

    Args:
        value: Raw cell value from openpyxl.  Must be castable to ``float``.

    Returns:
        Percentage as a whole-number ``float`` (e.g. ``42.0``).

    Raises:
        ValueError: If ``value`` is ``None``.
    """
    if value is None:
        raise ValueError("Encountered empty percentage cell")
    number = float(value)  # type: ignore[arg-type]
    pct = number * 100.0 if 0.0 <= number <= 1.0 else number
    return float(int(round(pct)))


def _infer_year_from_url(xlsx_url: str) -> int | None:
    """Extract the first four-digit year (20xx) found in a URL string.

    Args:
        xlsx_url: URL of the XLSX file, e.g.
            ``"https://example.com/media/voting-intention-10-feb-2025.xlsx"``.

    Returns:
        Integer year, or ``None`` if no ``20xx`` pattern is found.
    """
    matches = re.findall(r"(20\d{2})", xlsx_url)
    if not matches:
        return None
    return int(matches[0])


def _infer_fieldwork_from_source_url(
    source_url: str,
    *,
    default_year: int | None,
) -> tuple[date, date] | None:
    """Attempt to derive a fieldwork date from patterns embedded in the XLSX URL.

    Used as a fallback when no explicit fieldwork label is found in the
    workbook.  Looks for patterns such as
    ``voting-intention-february-10`` or ``voting-intention-february10``
    in the URL (case-insensitive).

    Args:
        source_url: URL of the XLSX file.
        default_year: Year to assign to the parsed date.  If ``None`` the
            function returns ``None`` immediately.

    Returns:
        Tuple of ``(fieldwork_start, fieldwork_end)`` with both elements set
        to the same single day, or ``None`` if no match is found.
    """
    if default_year is None:
        return None

    lower = source_url.lower()
    month_pattern = "|".join(
        [
            "january",
            "february",
            "march",
            "april",
            "may",
            "june",
            "july",
            "august",
            "september",
            "october",
            "november",
            "december",
        ]
    )

    patterns = [
        re.compile(rf"voting[-_ ]intention[-_ ]({month_pattern})[-_ ](\d{{1,2}})"),
        re.compile(rf"voting[-_ ]intention[-_ ]({month_pattern})(\d{{1,2}})"),
    ]

    for pattern in patterns:
        match = pattern.search(lower)
        if not match:
            continue
        month = _month_number(match.group(1))
        day = int(match.group(2))
        if month is None:
            continue
        parsed_day = date(default_year, month, day)
        return parsed_day, parsed_day

    return None


def _find_headline_sheet(workbook: Any) -> tuple[Any, int]:
    """Locate the headline voting-intention worksheet and its header row.

    First pass: finds a sheet containing both ``"All"`` and
    ``"East Midlands"`` column headers, prioritising sheets whose names
    suggest they are the headline voting-intention tab.

    Second pass (fallback): relaxes the regional requirement and accepts any
    sheet that has an ``"All"`` column alongside known party labels.

    Args:
        workbook: An openpyxl ``Workbook`` object.

    Returns:
        Tuple of ``(worksheet, header_row_number)`` where ``header_row_number``
        is the 1-based row index of the column-header row.

    Raises:
        ValueError: If no suitable sheet or header row can be found.
    """
    def _sheet_priority(sheet_name: str) -> tuple[int, str]:
        """Return a sort key that ranks sheets by name relevance.

        Args:
            sheet_name: Name of the worksheet to rank.

        Returns:
            Tuple of ``(priority_int, lowered_name)`` where a lower integer
            means higher priority.
        """
        lowered = sheet_name.lower()
        if "votingintention (headline)" in lowered:
            return (0, lowered)
        if "headline" in lowered and "votingintention" in lowered:
            return (1, lowered)
        if "votingintention" in lowered:
            return (2, lowered)
        if "corbyn" in lowered:
            return (3, lowered)
        return (9, lowered)

    for sheet_name in sorted(workbook.sheetnames, key=_sheet_priority):
        ws = workbook[sheet_name]
        for row in range(1, 35):
            values = [_cell_text(ws.cell(row, col).value) for col in range(1, 60)]
            if "All" in values and "East Midlands" in values:
                return ws, row

    for sheet_name in sorted(workbook.sheetnames, key=_sheet_priority):
        ws = workbook[sheet_name]
        for row in range(1, 35):
            values = [_cell_text(ws.cell(row, col).value) for col in range(1, 60)]
            if "All" not in values:
                continue
            if any(label in values for label in ("Conservative", "Labour", "Reform UK", "The Green Party")):
                return ws, row

    raise ValueError("Could not locate headline voting intention table")


def parse_poll(
    workbook: Any,
    *,
    source_url: str,
    fieldwork_year_hint: int | None = None,
) -> ParsedPoll:
    """Extract structured poll data from an openpyxl workbook.

    Searches the workbook for the headline voting-intention sheet and
    extracts sample size, fieldwork dates, and per-party/per-region vote
    share percentages.

    Fieldwork dates are resolved in priority order:
    1. ``"Fieldwork"`` label found in the cover page or first sheet.
    2. Month/day pattern embedded in ``source_url`` combined with the
       inferred or hinted year.

    Sample size is resolved in priority order:
    1. ``"Sample size"`` label found in the cover page or first sheet.
    2. ``"Unweighted n"`` or ``"Weighted n"`` row in the headline table.

    Args:
        workbook: An openpyxl ``Workbook`` object (``data_only=True``).
        source_url: URL the workbook was downloaded from; used for year and
            date inference when metadata labels are absent.
        fieldwork_year_hint: Override the year inferred from ``source_url``
            when the URL contains no year.  Optional.

    Returns:
        A :class:`ParsedPoll` containing sample size, fieldwork date range,
        and a mapping of canonical party name → region name → percentage.

    Raises:
        ValueError: If fieldwork dates, sample size, required region columns,
            or required party rows cannot be found.
    """
    default_year = fieldwork_year_hint if fieldwork_year_hint is not None else _infer_year_from_url(source_url)

    sheet, header_row = _find_headline_sheet(workbook)

    fieldwork_raw = None
    sample_raw = None
    preferred_sheet_names = ["Cover page", workbook.sheetnames[0]]
    for sheet_name in preferred_sheet_names + workbook.sheetnames:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        if fieldwork_raw is None:
            fieldwork_raw = _find_label_value(ws, "Fieldwork")
        if sample_raw is None:
            sample_raw = _find_label_value(ws, "Sample size")
        if fieldwork_raw and sample_raw:
            break

    if not fieldwork_raw:
        inferred = _infer_fieldwork_from_source_url(source_url, default_year=default_year)
        if inferred is None:
            raise ValueError("Fieldwork date not found in workbook")
        fieldwork_start, fieldwork_end = inferred
    else:
        fieldwork_start, fieldwork_end = parse_fieldwork(fieldwork_raw, default_year=default_year)

    if not sample_raw:
        for row in range(header_row + 1, header_row + 80):
            row_label = normalize_name(_cell_text(sheet.cell(row, 1).value))
            if row_label in {"unweighted n", "weighted n"}:
                sample_candidate = _cell_text(sheet.cell(row, 2).value)
                if sample_candidate:
                    sample_raw = sample_candidate
                    break
    if not sample_raw:
        raise ValueError("Sample size not found in workbook")
    sample_size = _as_int(sample_raw)

    population_label = ""
    for sheet_name in ["Cover page", workbook.sheetnames[0]] + workbook.sheetnames:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        value = _find_label_value(ws, "Population effectively represented")
        if value:
            population_label = value
            break

    region_columns: dict[int, str] = {}
    national_column: int | None = None
    for col in range(1, 120):
        header = _cell_text(sheet.cell(header_row, col).value)
        if header == "All":
            national_column = col
        mapped_region = REGION_HEADER_TO_INTERNAL.get(header)
        if mapped_region:
            region_columns[col] = mapped_region

    if national_column is None:
        raise ValueError("Could not locate 'All' (national) column in headline table")
    is_special_population = "16-17" in population_label.replace(" ", "")
    if len(region_columns) < 6 and not is_special_population:
        raise ValueError("Could not resolve sufficient region columns in headline table")

    party_region_percentages: dict[str, dict[str, float]] = {}
    for row in range(header_row + 1, header_row + 60):
        party_label = _cell_text(sheet.cell(row, 1).value)
        if not party_label:
            continue

        normalized_label = normalize_name(party_label)
        if normalized_label in {"weighted n", "unweighted n", "weight"}:
            break

        canonical_party = PARTY_NAME_MAP.get(party_label)
        if canonical_party is None:
            continue

        region_values: dict[str, float] = {NATIONAL_KEY: _to_percentage(sheet.cell(row, national_column).value)}
        for col, internal_region in region_columns.items():
            region_values[internal_region] = _to_percentage(sheet.cell(row, col).value)

        party_region_percentages[canonical_party] = region_values

    required_parties = {
        "Conservative",
        "Labour",
        "Liberal Democrats",
        "Reform UK",
        "Green",
        "Scottish National Party",
        "Plaid Cymru",
        "Other",
    }
    missing = sorted(required_parties - set(party_region_percentages.keys()))
    if missing:
        raise ValueError(f"Missing expected party rows in workbook: {missing}")

    return ParsedPoll(
        sample_size=sample_size,
        fieldwork_start=fieldwork_start,
        fieldwork_end=fieldwork_end,
        party_region_percentages=party_region_percentages,
    )


def _find_existing_poll(
    db: Database,
    pollster_id: int,
    map_id: int,
    parsed: ParsedPoll,
) -> Poll | None:
    """Look up an existing Poll row that matches the parsed poll's key fields.

    Matches on pollster, map, fieldwork start/end dates, and sample size.

    Args:
        db: Active :class:`Database` instance.
        pollster_id: Primary key of the pollster in the database.
        map_id: Primary key of the map in the database.
        parsed: Parsed poll data whose fieldwork dates and sample size are
            used as match criteria.

    Returns:
        The matching :class:`Poll` row, or ``None`` if no match exists.
    """
    with db.session() as session:
        return session.execute(
            select(Poll).where(
                Poll.pollster_id == pollster_id,
                Poll.map_id == map_id,
                Poll.fieldwork_start == parsed.fieldwork_start,
                Poll.fieldwork_end == parsed.fieldwork_end,
                Poll.sample_size == parsed.sample_size,
            )
        ).scalar_one_or_none()


def build_import_plan(
    db: Database,
    *,
    xlsx_url: str = DEFAULT_XLSX_URL,
    map_name: str = DEFAULT_MAP_NAME,
    pollster_identifier: str = DEFAULT_POLLSTER_IDENTIFIER,
    fieldwork_year_hint: int | None = None,
) -> ImportPlan:
    """Download and parse an XLSX poll, then build a full import plan.

    Downloads the workbook, parses poll metadata and vote shares, resolves
    database IDs for the map, regions, parties, and pollster, and assembles
    the complete set of :class:`PlannedPollRow` objects ready for insertion.

    No writes are made to the database; call :func:`commit_import_plan` to
    apply the plan.

    Args:
        db: Active :class:`Database` instance.
        xlsx_url: URL of the More in Common XLSX file to import.  Defaults
            to :data:`DEFAULT_XLSX_URL`.
        map_name: Name of the constituency map to associate with this poll.
            Must already exist in the database.  Defaults to
            :data:`DEFAULT_MAP_NAME`.
        pollster_identifier: Stable string identifier for the pollster row.
            Defaults to :data:`DEFAULT_POLLSTER_IDENTIFIER`.
        fieldwork_year_hint: Explicit year to use when the XLSX URL contains
            no four-digit year.  Optional.

    Returns:
        An :class:`ImportPlan` describing what would be created or updated.

    Raises:
        ValueError: If the map is not found, required parties are absent from
            the database, or the workbook cannot be parsed.
    """
    poll_map = db.get_map_by_name(map_name)
    if poll_map is None:
        raise ValueError(f"Map not found: {map_name!r}")

    regions = db.get_regions_for_map(poll_map.id)

    region_ids_by_name: dict[str, int] = {}
    for region in regions:
        region_ids_by_name[region.name] = region.id

    regions_mapping = "\n".join(
        f"{region_name}:{region_ids_by_name[region_name]}"
        for region_name in sorted(region_ids_by_name.keys())
    )

    workbook = extract_workbook(xlsx_url)
    parsed = parse_poll(
        workbook,
        source_url=xlsx_url,
        fieldwork_year_hint=fieldwork_year_hint,
    )

    pollster = db.get_pollster_by_identifier(pollster_identifier)
    pollster_exists = pollster is not None

    party_by_name = {party.name: party for party in db.get_all_parties()}
    missing_parties = [
        party_name
        for party_name in sorted(set(PARTY_NAME_MAP.values()))
        if party_name not in party_by_name
    ]
    if missing_parties:
        raise ValueError(
            "Missing parties in database (run party importer first): "
            f"{missing_parties}"
        )

    rows: list[PlannedPollRow] = []
    for party_name, region_values in parsed.party_region_percentages.items():
        party_id = party_by_name[party_name].id

        national_percentage = region_values.get(NATIONAL_KEY)
        if national_percentage is not None:
            rows.append(
                PlannedPollRow(
                    party_id=party_id,
                    party_name=party_name,
                    region_id=None,
                    region_name="National",
                    percentage=national_percentage,
                )
            )

        for region_name, region_id in region_ids_by_name.items():
            percentage = region_values.get(region_name, 0.0)
            rows.append(
                PlannedPollRow(
                    party_id=party_id,
                    party_name=party_name,
                    region_id=region_id,
                    region_name=region_name,
                    percentage=percentage,
                )
            )

    existing_poll = (
        _find_existing_poll(db, pollster.id, poll_map.id, parsed)
        if pollster is not None
        else None
    )

    return ImportPlan(
        pollster_identifier=pollster_identifier,
        pollster_name=(pollster.name if pollster else "More in Common"),
        pollster_id=(pollster.id if pollster else None),
        pollster_exists=pollster_exists,
        regions_mapping=regions_mapping,
        map_id=poll_map.id,
        map_name=poll_map.name,
        source_url=xlsx_url,
        parsed=parsed,
        poll_id=(existing_poll.id if existing_poll else None),
        poll_exists=existing_poll is not None,
        rows=rows,
    )


def commit_import_plan(
    db: Database,
    plan: ImportPlan,
    *,
    replace_rows: bool = False,
) -> PollImportResult:
    """Write an import plan to the database, creating or updating records as needed.

    Creates the pollster if it does not exist, or updates its
    ``regions_mapping`` if it has changed.  Creates the poll if it does not
    exist, or updates its ``source_url`` if it has changed.

    By default, if the poll already has rows the function skips insertion and
    sets ``skipped_existing_rows=True`` in the result.  Pass
    ``replace_rows=True`` to delete existing rows before inserting.

    Args:
        db: Active :class:`Database` instance.
        plan: Import plan produced by :func:`build_import_plan`.
        replace_rows: When ``True``, delete all existing :class:`PollRow`
            records for the poll before inserting new ones.  Defaults to
            ``False``.

    Returns:
        A :class:`PollImportResult` summarising what was created, replaced,
        or skipped.

    Raises:
        ValueError: If the pollster lookup fails unexpectedly during commit.
    """
    if plan.pollster_exists:
        pollster = db.get_pollster_by_identifier(plan.pollster_identifier)
        if pollster is None:
            raise ValueError("Pollster lookup failed during commit")
        with db.session() as session:
            db_pollster = session.get(Pollster, pollster.id)
            if db_pollster is not None and db_pollster.regions_mapping != plan.regions_mapping:
                db_pollster.regions_mapping = plan.regions_mapping
        pollster_id = pollster.id
        created_pollster = False
    else:
        created = db.add_pollster(
            name=plan.pollster_name,
            identifier=plan.pollster_identifier,
            weight=1.0,
            regions_mapping=plan.regions_mapping,
        )
        pollster_id = created.id
        created_pollster = True

    existing_poll = _find_existing_poll(db, pollster_id, plan.map_id, plan.parsed)
    if existing_poll is None:
        created_poll = db.add_poll(
            pollster_id=pollster_id,
            map_id=plan.map_id,
            fieldwork_start=plan.parsed.fieldwork_start,
            fieldwork_end=plan.parsed.fieldwork_end,
            sample_size=plan.parsed.sample_size,
            source_url=plan.source_url,
        )
        poll_id = created_poll.id
        created_poll_row = True
    else:
        with db.session() as session:
            db_poll = session.get(Poll, existing_poll.id)
            if db_poll is not None and db_poll.source_url != plan.source_url:
                db_poll.source_url = plan.source_url
        poll_id = existing_poll.id
        created_poll_row = False

    with db.session() as session:
        existing_rows = session.execute(
            select(PollRow).where(PollRow.poll_id == poll_id)
        ).scalars().all()

        if existing_rows and not replace_rows:
            return PollImportResult(
                created_pollster=created_pollster,
                created_poll=created_poll_row,
                poll_id=poll_id,
                inserted_rows=0,
                replaced_rows=0,
                skipped_existing_rows=True,
            )

        replaced_rows = 0
        if existing_rows and replace_rows:
            for existing_row in existing_rows:
                session.delete(existing_row)
                replaced_rows += 1

        for row in plan.rows:
            session.add(
                PollRow(
                    poll_id=poll_id,
                    region_id=row.region_id,
                    party_id=row.party_id,
                    percentage=row.percentage,
                )
            )

    return PollImportResult(
        created_pollster=created_pollster,
        created_poll=created_poll_row,
        poll_id=poll_id,
        inserted_rows=len(plan.rows),
        replaced_rows=replaced_rows,
        skipped_existing_rows=False,
    )


def _cli_preview(plan: ImportPlan) -> None:
    """Print a dry-run summary of an import plan to stdout.

    Outputs the parsed poll metadata, whether the pollster and poll already
    exist, and each row that would be inserted.  No database changes are made.

    Args:
        plan: Import plan to preview.
    """
    print(
        "Parsed poll: "
        f"fieldwork={plan.parsed.fieldwork_start} to {plan.parsed.fieldwork_end}, "
        f"sample={plan.parsed.sample_size}"
    )
    if plan.pollster_exists:
        print(f"pollster exists: {plan.pollster_identifier}")
    else:
        print(f"[dry-run] would create pollster: {plan.pollster_identifier}")

    if plan.poll_exists and plan.poll_id is not None:
        print(f"poll exists: {plan.poll_id}")
    else:
        print("[dry-run] would create poll")

    for row in plan.rows:
        print(
            "[dry-run] would insert row: "
            f"party={row.party_name}, region={row.region_name}, "
            f"region_id={row.region_id}, pct={row.percentage:.2f}"
        )


def main() -> None:
    """CLI entry point for importing a More in Common XLSX poll.

    Parses command-line arguments, builds an import plan, and either prints
    a dry-run preview or commits the plan to the database.

    Command-line arguments:
        --xlsx-url (str): URL of the More in Common XLSX file.
            Defaults to :data:`DEFAULT_XLSX_URL`.
        --map-name (str): Name of the constituency map in the database.
            Defaults to :data:`DEFAULT_MAP_NAME`.
        --pollster-identifier (str): Stable identifier for the pollster.
            Defaults to :data:`DEFAULT_POLLSTER_IDENTIFIER`.
        --fieldwork-year-hint (int): Optional year override used when the URL
            contains no four-digit year.
        --replace-rows: If set, delete existing poll rows before inserting.
        --dry-run: If set, print a preview and exit without writing to the DB.
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx-url", default=DEFAULT_XLSX_URL)
    parser.add_argument("--map-name", default=DEFAULT_MAP_NAME)
    parser.add_argument("--pollster-identifier", default=DEFAULT_POLLSTER_IDENTIFIER)
    parser.add_argument("--fieldwork-year-hint", type=int, default=None)
    parser.add_argument(
        "--replace-rows",
        action="store_true",
        help="Delete existing rows for the poll before inserting new ones",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    db = Database()
    print(f"Fetching XLSX: {args.xlsx_url}")

    plan = build_import_plan(
        db,
        xlsx_url=args.xlsx_url,
        map_name=args.map_name,
        pollster_identifier=args.pollster_identifier,
        fieldwork_year_hint=args.fieldwork_year_hint,
    )

    if args.dry_run:
        _cli_preview(plan)
        return

    result = commit_import_plan(db, plan, replace_rows=args.replace_rows)
    if result.created_pollster:
        print(f"created pollster: {args.pollster_identifier}")
    else:
        print(f"pollster exists: {args.pollster_identifier}")

    if result.created_poll:
        print(f"created poll: {result.poll_id}")
    else:
        print(f"poll exists: {result.poll_id}")

    if result.skipped_existing_rows:
        print(
            f"poll {result.poll_id} already has rows; "
            "use --replace-rows to overwrite"
        )
    else:
        if result.replaced_rows:
            print(f"deleted existing rows: {result.replaced_rows}")
        print(f"inserted poll rows: {result.inserted_rows}")


if __name__ == "__main__":
    main()
