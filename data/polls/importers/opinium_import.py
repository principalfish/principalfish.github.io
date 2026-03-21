#!/usr/bin/env python3
"""Import an Opinium poll directly from an XLSX URL."""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.request import urlopen

from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import select

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from db import Database
from models import Poll, PollRow, Pollster
from polls.importers.types import PollImportResult

DEFAULT_XLSX_URL = "https://www.opinium.com/wp-content/uploads/2026/02/Observer-VI-2026-02-04-Web-Data-Tables-1152.xlsx"
DEFAULT_MAP_NAME = "UK Constituencies post 2022"
DEFAULT_POLLSTER_IDENTIFIER = "opinium"

NATIONAL_KEY = "__national__"

MACRO_TO_INTERNAL_REGIONS = {
    "North": [
        "North East England",
        "North West England",
        "Yorkshire and The Humber",
    ],
    "Mids": ["East Midlands", "West Midlands"],
    "London": ["London"],
    "South": ["East of England", "South East England", "South West England"],
    "Wales": ["Wales"],
    "Scotland": ["Scotland"],
    "Northern Ireland": ["Northern Ireland"],
}

PARTY_NAME_MAP = {
    "Con": "Conservative",
    "Conservative": "Conservative",
    "Lab": "Labour",
    "Labour": "Labour",
    "Lib Dem": "Liberal Democrats",
    "Liberal Democrat": "Liberal Democrats",
    "Liberal Democrats": "Liberal Democrats",
    "Reform": "Reform UK",
    "Reform UK": "Reform UK",
    "Green": "Green",
    "Other": "Other",
}


class ParsedPoll(BaseModel):
    """Parsed poll data extracted from source."""

    sample_size: int = Field(gt=0)
    fieldwork_start: date
    fieldwork_end: date
    party_macro_percentages: dict[str, dict[str, float]]


class PlannedPollRow(BaseModel):
    """A single poll row planned for DB insertion."""

    party_id: int
    party_name: str
    region_id: int | None
    region_name: str
    percentage: float = Field(ge=0, le=100)


class ImportPlan(BaseModel):
    """Full import plan: pollster, poll metadata, and rows."""

    pollster_identifier: str
    pollster_name: str
    pollster_id: int | None
    pollster_exists: bool
    regions_mapping: str
    map_id: int
    map_name: str
    source_url: str
    parsed: ParsedPoll
    poll_id: int | None
    poll_exists: bool
    rows: list[PlannedPollRow]


def _month_number(month_text: str) -> int | None:
    """Convert a month name or abbreviation to its integer number (1–12).

    Args:
        month_text: Month name or abbreviation, e.g. ``"Jan"``, ``"February"``,
            ``"Sept."``. Case-insensitive; trailing periods are stripped.

    Returns:
        Integer month number (1–12), or ``None`` if the text is not recognised.
    """
    month_map = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }
    return month_map.get(month_text.strip().lower().rstrip("."))


def _parse_fieldwork(value: str, default_year: int | None = None) -> tuple[date, date]:
    """Parse a fieldwork date-range string into a (start, end) pair of dates.

    Handles formats such as ``"3-5 February 2026"`` or ``"3-5 Feb"`` (when a
    ``default_year`` is supplied).  En-dashes and em-dashes are normalised to
    hyphens; ordinal suffixes (``st``, ``nd``, ``rd``, ``th``) are stripped
    before matching.

    Args:
        value: Raw fieldwork string as it appears in the spreadsheet.
        default_year: Year to use when no explicit 4-digit year is present in
            ``value``. Optional; must be provided for year-less strings or a
            ``ValueError`` is raised.

    Returns:
        A ``(fieldwork_start, fieldwork_end)`` tuple of :class:`datetime.date`
        objects.

    Raises:
        ValueError: If the string cannot be parsed or the month name is not
            recognised.
    """
    normalized = re.sub(r"\s+", " ", value.strip().replace("–", "-").replace("—", "-"))
    normalized = re.sub(r"(\d)(st|nd|rd|th)", r"\1", normalized, flags=re.IGNORECASE)

    range_with_year = re.compile(
        r"(\d{1,2})\s*-\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})"
    )
    range_no_year = re.compile(r"(\d{1,2})\s*-\s*(\d{1,2})\s+([A-Za-z]+)")

    match = range_with_year.search(normalized)
    if match:
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        year = int(match.group(4))
        if month is None:
            raise ValueError(f"Could not parse fieldwork string: {value!r}")
        return date(year, month, day_start), date(year, month, day_end)

    match = range_no_year.search(normalized)
    if match and default_year is not None:
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        if month is None:
            raise ValueError(f"Could not parse fieldwork string: {value!r}")
        return date(default_year, month, day_start), date(default_year, month, day_end)

    raise ValueError(f"Could not parse fieldwork string: {value!r}")


def _infer_year_from_url(xlsx_url: str, fallback: int | None = None) -> int:
    """Extract a four-digit year from an XLSX URL.

    Tries a path-segment match first (``/2026/``), then falls back to the
    first ``20XX`` sequence found anywhere in the URL.  If neither succeeds
    and a ``fallback`` is provided, that value is returned; otherwise raises.

    Args:
        xlsx_url: Full URL to the XLSX file.
        fallback: Year to return when the URL contains no recognisable year.
            Optional.

    Returns:
        Integer year inferred from the URL, or ``fallback`` if supplied.

    Raises:
        ValueError: If no year can be found and ``fallback`` is ``None``.
    """
    match = re.search(r"/(20\d{2})/", xlsx_url)
    if match:
        return int(match.group(1))
    matches = re.findall(r"(20\d{2})", xlsx_url)
    if matches:
        return int(matches[0])
    if fallback is not None:
        return fallback
    raise ValueError("Could not infer year from URL")


def extract_workbook(xlsx_url: str) -> Any:
    """Download an XLSX file from a URL and return the loaded openpyxl workbook.

    Args:
        xlsx_url: Publicly accessible URL pointing to an ``.xlsx`` file.

    Returns:
        An :class:`openpyxl.Workbook` instance loaded in data-only mode
        (formula results, not formulas).
    """
    payload = urlopen(xlsx_url).read()
    return load_workbook(filename=BytesIO(payload), data_only=True)


def _cell_text(value: object) -> str:
    """Return the string representation of a cell value, stripped of whitespace.

    Args:
        value: Raw cell value from openpyxl (any type, or ``None``).

    Returns:
        Stripped string, or ``""`` when ``value`` is ``None``.
    """
    if value is None:
        return ""
    return str(value).strip()


def _to_percentage(value: object) -> float:
    """Convert a cell value to a whole-number percentage float.

    Values in the range ``[0.0, 1.0]`` are assumed to be proportions and are
    multiplied by 100.  Values already expressed as percentages (``> 1.0``) are
    used directly.  The result is rounded to the nearest integer and returned as
    a ``float``.

    Args:
        value: Raw cell value, expected to be numeric.

    Returns:
        Percentage as a whole-number float (e.g. ``42.0``).

    Raises:
        ValueError: If ``value`` is ``None``.
    """
    if value is None:
        raise ValueError("Encountered empty percentage cell")
    number = float(value)  # type: ignore[arg-type]
    pct = number * 100.0 if 0.0 <= number <= 1.0 else number
    return float(int(round(pct)))


def _find_fieldwork_and_sample(workbook: Any, default_year: int) -> tuple[date, date, int | None]:
    """Scan the FRONT PAGE sheet for fieldwork dates and sample size.

    Reads up to the first 50 rows of column C (label) and column F (value),
    looking for cells whose labels contain ``"field date"``/``"field dates"``
    or ``"sample"``.  Falls back to the first sheet if no ``"FRONT PAGE"``
    sheet exists.

    Args:
        workbook: Loaded openpyxl workbook.
        default_year: Year passed to :func:`_parse_fieldwork` when the
            fieldwork string contains no explicit year.

    Returns:
        A ``(fieldwork_start, fieldwork_end, sample_size)`` tuple.
        ``sample_size`` is ``None`` when a sample value cannot be extracted.

    Raises:
        ValueError: If the fieldwork date row is not found, or if the date
            string cannot be parsed.
    """
    front = workbook["FRONT PAGE"] if "FRONT PAGE" in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    fieldwork_raw = None
    sample_raw = None
    for row in range(1, 50):
        label = _cell_text(front.cell(row, 3).value).lower()
        value = _cell_text(front.cell(row, 6).value)
        if not label and not value:
            continue
        if "field date" in label or "field dates" in label:
            fieldwork_raw = value
        if "sample" in label:
            sample_raw = value

    if not fieldwork_raw:
        raise ValueError("Fieldwork date not found in workbook")

    fieldwork_start, fieldwork_end = _parse_fieldwork(fieldwork_raw, default_year=default_year)

    sample_size = None
    if sample_raw:
        digits = re.sub(r"[^0-9]", "", sample_raw)
        if digits:
            sample_size = int(digits)

    return fieldwork_start, fieldwork_end, sample_size


def _find_headline_sheet(workbook: Any) -> Any:
    """Return the HeadlineVI worksheet from the workbook.

    Matches a sheet named exactly ``"headlinevi"`` (case-insensitive) or any
    sheet whose name contains ``"headline"``.

    Args:
        workbook: Loaded openpyxl workbook.

    Returns:
        The matching :class:`openpyxl.worksheet.worksheet.Worksheet`.

    Raises:
        ValueError: If no matching sheet is found.
    """
    for name in workbook.sheetnames:
        if name.lower() == "headlinevi" or "headline" in name.lower():
            return workbook[name]
    raise ValueError("Could not find HeadlineVI sheet")


def parse_poll(workbook: Any, *, source_url: str, fieldwork_year_hint: int | None = None) -> ParsedPoll:
    """Extract poll metadata and regional vote-share data from a loaded workbook.

    Reads fieldwork dates and sample size from the FRONT PAGE sheet, then
    parses party vote shares by macro-region from the HeadlineVI sheet.
    Optional parties (SNP, Plaid Cymru) and optional regions (Northern Ireland)
    are zero-filled when absent from the sheet.

    Args:
        workbook: Loaded openpyxl workbook, typically from
            :func:`extract_workbook`.
        source_url: URL the workbook was fetched from; used to infer the
            fieldwork year when not present in the date string.
        fieldwork_year_hint: Explicit year to fall back to if the year cannot
            be inferred from ``source_url``.  Optional.

    Returns:
        A :class:`ParsedPoll` containing sample size, fieldwork dates, and a
        ``party_macro_percentages`` mapping of
        ``{party_name: {macro_region_or_NATIONAL_KEY: percentage}}``.

    Raises:
        ValueError: If required sheets, headers, weighted-base row, or party
            rows are missing, or if any date/number cannot be parsed.
    """
    year = _infer_year_from_url(source_url, fallback=fieldwork_year_hint)
    fieldwork_start, fieldwork_end, sample_from_front = _find_fieldwork_and_sample(workbook, year)

    sheet = _find_headline_sheet(workbook)

    header_row = None
    for row in range(1, 20):
        first = _cell_text(sheet.cell(row, 1).value)
        second = _cell_text(sheet.cell(row, 2).value)
        if first == "" and second.lower() == "total":
            header_row = row
            break
    if header_row is None:
        raise ValueError("Could not locate headline header row")

    base_row = None
    for row in range(header_row + 1, header_row + 8):
        label = _cell_text(sheet.cell(row, 1).value).lower()
        if "base:" in label and "weighted" in label:
            base_row = row
            break
    if base_row is None:
        raise ValueError("Could not locate weighted base row in HeadlineVI sheet")

    weighted_base_text = _cell_text(sheet.cell(base_row, 2).value)
    digits = re.sub(r"[^0-9]", "", weighted_base_text)
    sample_size = int(digits) if digits else sample_from_front
    if sample_size is None:
        raise ValueError("Could not determine sample size")

    header_labels: dict[str, int] = {}
    for col in range(1, 60):
        header = _cell_text(sheet.cell(header_row, col).value)
        if header:
            header_labels[header] = col

    required_macro_headers = ["North", "Mids", "London", "South", "Wales", "Scotland"]
    for macro in required_macro_headers:
        if macro not in header_labels:
            raise ValueError(f"Missing expected regional header {macro!r} in HeadlineVI sheet")

    optional_macro_headers = ["Northern Ireland"]

    if "Total" not in header_labels:
        raise ValueError("Missing expected 'Total' header in HeadlineVI sheet")

    party_macro_percentages: dict[str, dict[str, float]] = {}
    for row in range(base_row + 1, base_row + 30):
        party_label = _cell_text(sheet.cell(row, 1).value)
        if not party_label:
            continue
        if party_label.lower().startswith("return to index"):
            break
        canonical = PARTY_NAME_MAP.get(party_label)
        if canonical is None:
            continue
        macro_values = {NATIONAL_KEY: _to_percentage(sheet.cell(row, header_labels["Total"]).value)}
        for macro in required_macro_headers:
            macro_values[macro] = _to_percentage(sheet.cell(row, header_labels[macro]).value)
        for macro in optional_macro_headers:
            if macro in header_labels:
                macro_values[macro] = _to_percentage(sheet.cell(row, header_labels[macro]).value)
            else:
                macro_values[macro] = 0.0
        party_macro_percentages[canonical] = macro_values

    for optional_party in ["Scottish National Party", "Plaid Cymru", "Other"]:
        if optional_party not in party_macro_percentages:
            party_macro_percentages[optional_party] = {NATIONAL_KEY: 0.0}
            for macro in required_macro_headers + optional_macro_headers:
                party_macro_percentages[optional_party][macro] = 0.0

    required_parties = {
        "Conservative",
        "Labour",
        "Liberal Democrats",
        "Reform UK",
        "Green",
        "Scottish National Party",
        "Plaid Cymru",
        "Other",
    }
    missing = sorted(required_parties - set(party_macro_percentages.keys()))
    if missing:
        raise ValueError(f"Missing expected party rows in workbook: {missing}")

    return ParsedPoll(
        sample_size=sample_size,
        fieldwork_start=fieldwork_start,
        fieldwork_end=fieldwork_end,
        party_macro_percentages=party_macro_percentages,
    )


def _find_existing_poll(db: Database, pollster_id: int, map_id: int, parsed: ParsedPoll) -> Poll | None:
    """Query the database for a poll matching the given pollster, map, and fieldwork details.

    Matches on pollster ID, map ID, fieldwork start/end dates, and sample size.

    Args:
        db: Active :class:`~db.Database` connection.
        pollster_id: Primary key of the pollster record.
        map_id: Primary key of the constituency map record.
        parsed: Parsed poll data supplying the fieldwork dates and sample size
            to match against.

    Returns:
        The matching :class:`~models.Poll` row, or ``None`` if no match exists.
    """
    with db.session() as session:
        return session.execute(
            select(Poll).where(
                Poll.pollster_id == pollster_id,
                Poll.map_id == map_id,
                Poll.fieldwork_start == parsed.fieldwork_start,
                Poll.fieldwork_end == parsed.fieldwork_end,
                Poll.sample_size == parsed.sample_size,
            )
        ).scalar_one_or_none()


def build_import_plan(
    db: Database,
    *,
    xlsx_url: str = DEFAULT_XLSX_URL,
    map_name: str = DEFAULT_MAP_NAME,
    pollster_identifier: str = DEFAULT_POLLSTER_IDENTIFIER,
    fieldwork_year_hint: int | None = None,
) -> ImportPlan:
    """Download the Opinium XLSX, parse it, and build a complete import plan.

    Resolves all DB references (map, regions, pollster, parties) and expands
    macro-region vote shares into individual per-region :class:`PlannedPollRow`
    objects.  Does **not** write anything to the database.

    Args:
        db: Active :class:`~db.Database` connection.
        xlsx_url: URL of the Opinium XLSX data-tables file.  Defaults to
            :data:`DEFAULT_XLSX_URL`.
        map_name: Name of the constituency map to associate with the poll.
            Must already exist in the database.  Defaults to
            :data:`DEFAULT_MAP_NAME`.
        pollster_identifier: Short identifier used to look up the pollster
            record (e.g. ``"opinium"``).  Defaults to
            :data:`DEFAULT_POLLSTER_IDENTIFIER`.
        fieldwork_year_hint: Fallback year for fieldwork-date parsing when the
            year cannot be inferred from ``xlsx_url``.  Optional.

    Returns:
        A fully populated :class:`ImportPlan` ready to be passed to
        :func:`commit_import_plan`.

    Raises:
        ValueError: If the map or any required region/party is missing from
            the database, or if the workbook cannot be parsed.
    """
    poll_map = db.get_map_by_name(map_name)
    if poll_map is None:
        raise ValueError(f"Map not found: {map_name!r}")

    regions = db.get_regions_for_map(poll_map.id)
    regions_by_name = {region.name: region for region in regions}

    macro_to_region_ids: dict[str, list[int]] = {}
    for macro_name, internal_names in MACRO_TO_INTERNAL_REGIONS.items():
        ids: list[int] = []
        for internal_name in internal_names:
            region = regions_by_name.get(internal_name)
            if region is None:
                raise ValueError(f"Region {internal_name!r} not found in map {map_name!r}")
            ids.append(region.id)
        macro_to_region_ids[macro_name] = ids

    regions_mapping = "\n".join(
        f"{macro}:{','.join(str(region_id) for region_id in region_ids)}"
        for macro, region_ids in macro_to_region_ids.items()
    )

    workbook = extract_workbook(xlsx_url)
    parsed = parse_poll(workbook, source_url=xlsx_url, fieldwork_year_hint=fieldwork_year_hint)

    pollster = db.get_pollster_by_identifier(pollster_identifier)
    pollster_exists = pollster is not None

    party_by_name = {party.name: party for party in db.get_all_parties()}
    missing_parties = [
        party_name
        for party_name in sorted(set(PARTY_NAME_MAP.values()))
        if party_name not in party_by_name
    ]
    if missing_parties:
        raise ValueError(
            "Missing parties in database (run party importer first): "
            f"{missing_parties}"
        )

    rows: list[PlannedPollRow] = []
    for party_name, macro_values in parsed.party_macro_percentages.items():
        national_percentage = macro_values.get(NATIONAL_KEY, 0.0)
        rows.append(
            PlannedPollRow(
                party_id=party_by_name[party_name].id,
                party_name=party_name,
                region_id=None,
                region_name="National",
                percentage=national_percentage,
            )
        )

        for macro_name, region_ids in macro_to_region_ids.items():
            percentage = macro_values.get(macro_name, 0.0)
            for region_id in region_ids:
                region_name = next(region.name for region in regions if region.id == region_id)
                rows.append(
                    PlannedPollRow(
                        party_id=party_by_name[party_name].id,
                        party_name=party_name,
                        region_id=region_id,
                        region_name=region_name,
                        percentage=percentage,
                    )
                )

    existing_poll = (
        _find_existing_poll(db, pollster.id, poll_map.id, parsed)
        if pollster is not None
        else None
    )

    return ImportPlan(
        pollster_identifier=pollster_identifier,
        pollster_name=(pollster.name if pollster else "Opinium"),
        pollster_id=(pollster.id if pollster else None),
        pollster_exists=pollster_exists,
        regions_mapping=regions_mapping,
        map_id=poll_map.id,
        map_name=poll_map.name,
        source_url=xlsx_url,
        parsed=parsed,
        poll_id=(existing_poll.id if existing_poll else None),
        poll_exists=existing_poll is not None,
        rows=rows,
    )


def commit_import_plan(
    db: Database,
    plan: ImportPlan,
    *,
    replace_rows: bool = False,
) -> PollImportResult:
    """Persist an import plan to the database.

    Creates the pollster record if it does not yet exist; otherwise updates its
    ``regions_mapping`` if it has changed.  Creates the poll record if absent;
    otherwise updates ``source_url`` if it has changed.  Inserts
    :class:`~models.PollRow` records for all rows in the plan unless rows
    already exist and ``replace_rows`` is ``False``.

    Args:
        db: Active :class:`~db.Database` connection.
        plan: Fully built :class:`ImportPlan` from :func:`build_import_plan`.
        replace_rows: When ``True``, delete any existing poll rows before
            inserting the new ones.  When ``False`` (the default), skip
            insertion if rows already exist.

    Returns:
        A :class:`~polls.importers.types.PollImportResult` summarising what was
        created, replaced, inserted, or skipped.

    Raises:
        ValueError: If the pollster lookup fails unexpectedly during commit.
    """
    if plan.pollster_exists:
        pollster = db.get_pollster_by_identifier(plan.pollster_identifier)
        if pollster is None:
            raise ValueError("Pollster lookup failed during commit")
        with db.session() as session:
            db_pollster = session.get(Pollster, pollster.id)
            if db_pollster is not None and db_pollster.regions_mapping != plan.regions_mapping:
                db_pollster.regions_mapping = plan.regions_mapping
        pollster_id = pollster.id
        created_pollster = False
    else:
        created = db.add_pollster(
            name=plan.pollster_name,
            identifier=plan.pollster_identifier,
            weight=1.0,
            regions_mapping=plan.regions_mapping,
        )
        pollster_id = created.id
        created_pollster = True

    existing_poll = _find_existing_poll(db, pollster_id, plan.map_id, plan.parsed)
    if existing_poll is None:
        created_poll = db.add_poll(
            pollster_id=pollster_id,
            map_id=plan.map_id,
            fieldwork_start=plan.parsed.fieldwork_start,
            fieldwork_end=plan.parsed.fieldwork_end,
            sample_size=plan.parsed.sample_size,
            source_url=plan.source_url,
        )
        poll_id = created_poll.id
        created_poll_row = True
    else:
        with db.session() as session:
            db_poll = session.get(Poll, existing_poll.id)
            if db_poll is not None and db_poll.source_url != plan.source_url:
                db_poll.source_url = plan.source_url
        poll_id = existing_poll.id
        created_poll_row = False

    with db.session() as session:
        existing_rows = session.execute(
            select(PollRow).where(PollRow.poll_id == poll_id)
        ).scalars().all()

        if existing_rows and not replace_rows:
            return PollImportResult(
                created_pollster=created_pollster,
                created_poll=created_poll_row,
                poll_id=poll_id,
                inserted_rows=0,
                replaced_rows=0,
                skipped_existing_rows=True,
            )

        replaced_rows = 0
        if existing_rows and replace_rows:
            for existing_row in existing_rows:
                session.delete(existing_row)
                replaced_rows += 1

        for row in plan.rows:
            session.add(
                PollRow(
                    poll_id=poll_id,
                    region_id=row.region_id,
                    party_id=row.party_id,
                    percentage=row.percentage,
                )
            )

    return PollImportResult(
        created_pollster=created_pollster,
        created_poll=created_poll_row,
        poll_id=poll_id,
        inserted_rows=len(plan.rows),
        replaced_rows=replaced_rows,
        skipped_existing_rows=False,
    )


def _cli_preview(plan: ImportPlan) -> None:
    """Print a dry-run preview of what the import plan would create or insert.

    Outputs parsed fieldwork dates and sample size, whether the pollster and
    poll records already exist, and one line per planned poll row.

    Args:
        plan: Import plan to preview, as returned by :func:`build_import_plan`.
    """
    print(
        "Parsed poll: "
        f"fieldwork={plan.parsed.fieldwork_start} to {plan.parsed.fieldwork_end}, "
        f"sample={plan.parsed.sample_size}"
    )
    if plan.pollster_exists:
        print(f"pollster exists: {plan.pollster_identifier}")
    else:
        print(f"[dry-run] would create pollster: {plan.pollster_identifier}")

    if plan.poll_exists and plan.poll_id is not None:
        print(f"poll exists: {plan.poll_id}")
    else:
        print("[dry-run] would create poll")

    for row in plan.rows:
        print(
            "[dry-run] would insert row: "
            f"party={row.party_name}, region={row.region_name}, "
            f"region_id={row.region_id}, pct={row.percentage:.2f}"
        )


def main() -> None:
    """CLI entry point for importing an Opinium poll from an XLSX URL.

    Parses command-line arguments, builds an import plan by fetching and
    parsing the specified XLSX file, then either prints a dry-run preview
    (``--dry-run``) or commits the plan to the database.

    Command-line arguments:
        --xlsx-url: URL of the Opinium XLSX data-tables file.
            Default: ``DEFAULT_XLSX_URL``.
        --map-name: Name of the constituency map in the database.
            Default: ``DEFAULT_MAP_NAME``.
        --pollster-identifier: Short identifier for the pollster record.
            Default: ``DEFAULT_POLLSTER_IDENTIFIER``.
        --fieldwork-year-hint: Integer year to use when the fieldwork year
            cannot be inferred from the URL.  Optional.
        --replace-rows: Flag; when set, deletes existing poll rows before
            inserting new ones.
        --dry-run: Flag; when set, prints a preview without writing to the DB.
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx-url", default=DEFAULT_XLSX_URL)
    parser.add_argument("--map-name", default=DEFAULT_MAP_NAME)
    parser.add_argument("--pollster-identifier", default=DEFAULT_POLLSTER_IDENTIFIER)
    parser.add_argument("--fieldwork-year-hint", type=int, default=None)
    parser.add_argument(
        "--replace-rows",
        action="store_true",
        help="Delete existing rows for the poll before inserting new ones",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    db = Database()
    print(f"Fetching XLSX: {args.xlsx_url}")

    plan = build_import_plan(
        db,
        xlsx_url=args.xlsx_url,
        map_name=args.map_name,
        pollster_identifier=args.pollster_identifier,
        fieldwork_year_hint=args.fieldwork_year_hint,
    )

    if args.dry_run:
        _cli_preview(plan)
        return

    result = commit_import_plan(db, plan, replace_rows=args.replace_rows)
    if result.created_pollster:
        print(f"created pollster: {args.pollster_identifier}")
    else:
        print(f"pollster exists: {args.pollster_identifier}")

    if result.created_poll:
        print(f"created poll: {result.poll_id}")
    else:
        print(f"poll exists: {result.poll_id}")

    if result.skipped_existing_rows:
        print(
            f"poll {result.poll_id} already has rows; "
            "use --replace-rows to overwrite"
        )
    else:
        if result.replaced_rows:
            print(f"deleted existing rows: {result.replaced_rows}")
        print(f"inserted poll rows: {result.inserted_rows}")


if __name__ == "__main__":
    main()
