#!/usr/bin/env python3
"""Import a Survation poll directly from an XLSX URL."""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import select

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from db import Database
from models import Poll, PollRow, Pollster
from polls.import_types import PollImportResult

DEFAULT_XLSX_URL = "https://cdn.survation.com/wp-content/uploads/2026/01/30163059/Survation_Voter_Intention_Jan_2026.xlsx"
DEFAULT_MAP_NAME = "UK Constituencies post 2022"
DEFAULT_POLLSTER_IDENTIFIER = "survation"
NATIONAL_KEY = "__national__"

SOURCE_REGION_TO_INTERNAL = {
    "East Midlands": "East Midlands",
    "East of England": "East of England",
    "London": "London",
    "North East": "North East England",
    "North West": "North West England",
    "South East": "South East England",
    "South West": "South West England",
    "West Midlands": "West Midlands",
    "Yorkshire and The Humber": "Yorkshire and The Humber",
    "Scotland": "Scotland",
    "Wales": "Wales",
    "Northern Ireland": "Northern Ireland",
}

PARTY_NAME_MAP = {
    "Conservative": "Conservative",
    "Labour": "Labour",
    "Liberal Democrats": "Liberal Democrats",
    "Liberal Democrat": "Liberal Democrats",
    "Reform UK": "Reform UK",
    "Green": "Green",
    "Green Party": "Green",
    "Scottish National Party": "Scottish National Party",
    "Scottish National Party (SNP)": "Scottish National Party",
    "SNP": "Scottish National Party",
    "Plaid Cymru": "Plaid Cymru",
    "Another party": "Other",
    "Other": "Other",
}


class ParsedPoll(BaseModel):
    """Parsed poll data extracted from source."""

    sample_size: int = Field(gt=0)
    fieldwork_start: date
    fieldwork_end: date
    party_region_percentages: dict[str, dict[str, float]]


class PlannedPollRow(BaseModel):
    """A single poll row planned for DB insertion."""

    party_id: int
    party_name: str
    region_id: int | None
    region_name: str
    percentage: float = Field(ge=0, le=100)


class ImportPlan(BaseModel):
    """Full import plan: pollster, poll metadata, and rows."""

    pollster_identifier: str
    pollster_name: str
    pollster_id: int | None
    pollster_exists: bool
    regions_mapping: str
    map_id: int
    map_name: str
    source_url: str
    parsed: ParsedPoll
    poll_id: int | None
    poll_exists: bool
    rows: list[PlannedPollRow]


def _month_number(month_text: str) -> int | None:
    month_map = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }
    return month_map.get(month_text.strip().lower().rstrip("."))


def _infer_year(url: str, fallback: int | None = None) -> int:
    match = re.search(r"/(20\d{2})/", url)
    if match:
        return int(match.group(1))
    matches = re.findall(r"(20\d{2})", url)
    if matches:
        return int(matches[0])
    if fallback is not None:
        return fallback
    raise ValueError("Could not infer year from URL")


def _parse_fieldwork(fieldwork_text: str, default_year: int | None = None) -> tuple[date, date]:
    normalized = re.sub(r"\s+", " ", fieldwork_text.strip().replace("–", "-").replace("—", "-"))
    normalized = re.sub(r"(\d)(st|nd|rd|th)", r"\1", normalized, flags=re.IGNORECASE)

    pattern_same_month = re.compile(r"(\d{1,2})\s*-\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})")
    match = pattern_same_month.search(normalized)
    if match:
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        year = int(match.group(4))
        if month is None:
            raise ValueError(f"Could not parse month in fieldwork: {fieldwork_text!r}")
        return date(year, month, day_start), date(year, month, day_end)

    pattern_no_year = re.compile(r"(\d{1,2})\s*-\s*(\d{1,2})\s+([A-Za-z]+)")
    match = pattern_no_year.search(normalized)
    if match and default_year is not None:
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        if month is None:
            raise ValueError(f"Could not parse month in fieldwork: {fieldwork_text!r}")
        return date(default_year, month, day_start), date(default_year, month, day_end)

    pattern_cross_no_year = re.compile(r"(\d{1,2})\s+([A-Za-z]+)\s*-\s*(\d{1,2})\s+([A-Za-z]+)")
    match = pattern_cross_no_year.search(normalized)
    if match and default_year is not None:
        day_start = int(match.group(1))
        month_start = _month_number(match.group(2))
        day_end = int(match.group(3))
        month_end = _month_number(match.group(4))
        if month_start is None or month_end is None:
            raise ValueError(f"Could not parse months in fieldwork: {fieldwork_text!r}")
        year_start = default_year - 1 if month_start > month_end else default_year
        return date(year_start, month_start, day_start), date(default_year, month_end, day_end)

    pattern_cross_month = re.compile(
        r"(\d{1,2})\s+([A-Za-z]+)\s*-\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})"
    )
    match = pattern_cross_month.search(normalized)
    if match:
        day_start = int(match.group(1))
        month_start = _month_number(match.group(2))
        day_end = int(match.group(3))
        month_end = _month_number(match.group(4))
        year_end = int(match.group(5))
        if month_start is None or month_end is None:
            raise ValueError(f"Could not parse months in fieldwork: {fieldwork_text!r}")
        year_start = year_end - 1 if month_start > month_end else year_end
        return date(year_start, month_start, day_start), date(year_end, month_end, day_end)

    raise ValueError(f"Could not parse fieldwork string: {fieldwork_text!r}")


def extract_workbook(xlsx_url: str) -> Any:
    req = Request(xlsx_url, headers={"User-Agent": "Mozilla/5.0 (compatible; poll-importer/1.0)"})
    with urlopen(req, timeout=50) as response:
        payload = response.read()

    if not payload.startswith(b"PK"):
        raise ValueError(f"Could not fetch XLSX payload from URL: {xlsx_url}")

    return load_workbook(filename=BytesIO(payload), data_only=True)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_percentage(value: object) -> float:
    if value is None:
        raise ValueError("Encountered empty percentage cell")
    number = float(value)  # type: ignore[arg-type]
    pct = number * 100.0 if 0.0 <= number <= 1.0 else number
    return float(int(round(pct)))


def _to_percentage_or_zero(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, str):
        cleaned = value.strip()
        if cleaned in {"", "-", "–", "—", "- ", "-\u00a0"}:
            return 0.0
    return _to_percentage(value)


def _cover_sheet(workbook: Any) -> Any:
    for name in workbook.sheetnames:
        if "cover" in name.lower() and "method" in name.lower():
            return workbook[name]
    return workbook[workbook.sheetnames[0]]


def _parse_cover_metadata(workbook: Any, default_year: int) -> tuple[date, date, int]:
    cover = _cover_sheet(workbook)

    fieldwork_raw = None
    sample_size = None

    for row in range(1, 80):
        label = _cell_text(cover.cell(row, 1).value).lower()
        if "fieldwork dates" in label:
            fieldwork_raw = _cell_text(cover.cell(row + 1, 1).value)
        if label == "sample size":
            candidate = cover.cell(row + 1, 1).value
            if isinstance(candidate, (int, float)):
                sample_size = int(round(float(candidate)))
            else:
                digits = re.sub(r"[^0-9]", "", _cell_text(candidate))
                if digits:
                    sample_size = int(digits)

    if not fieldwork_raw:
        raise ValueError("Fieldwork dates not found in cover sheet")
    if sample_size is None:
        raise ValueError("Sample size not found in cover sheet")

    fieldwork_start, fieldwork_end = _parse_fieldwork(fieldwork_raw, default_year=default_year)
    return fieldwork_start, fieldwork_end, sample_size


def _tables_sheet(workbook: Any) -> Any:
    for name in workbook.sheetnames:
        lowered = name.strip().lower()
        if lowered in {"tables", "table"}:
            return workbook[name]
    for name in workbook.sheetnames:
        lowered = name.strip().lower()
        if lowered.startswith("tables"):
            return workbook[name]
    for name in workbook.sheetnames:
        lowered = name.strip().lower()
        if "table" in lowered and "contents" not in lowered:
            return workbook[name]
    raise ValueError("Could not find tables sheet")


def _find_vi_table_start(sheet: Any) -> int:
    starts: list[int] = []
    preferred: list[int] = []

    for row in range(1, sheet.max_row + 1):
        header = _cell_text(sheet.cell(row, 1).value)
        header_lower = header.lower()
        if "table_" not in header_lower:
            continue
        if "general election" not in header_lower and "westminster election" not in header_lower:
            continue
        if "for which party" not in header_lower or "vote" not in header_lower:
            continue

        starts.append(row)
        base = _cell_text(sheet.cell(row + 1, 1).value).lower()
        if "undecided" in base and "remove" in base:
            preferred.append(row)

    if preferred:
        return preferred[-1]
    if starts:
        return starts[-1]
    raise ValueError("Could not locate voting intention table in Survation workbook")


def _parse_party_region_percentages(workbook: Any) -> dict[str, dict[str, float]]:
    sheet = _tables_sheet(workbook)
    start_row = _find_vi_table_start(sheet)

    region_header_row = None
    for row in range(start_row + 1, min(sheet.max_row, start_row + 12)):
        headers = {_cell_text(sheet.cell(row, col).value) for col in range(11, 40)}
        if any(header in SOURCE_REGION_TO_INTERNAL for header in headers):
            region_header_row = row
            break

    region_columns: dict[str, int] = {}
    if region_header_row is not None:
        for col in range(11, 45):
            header = _cell_text(sheet.cell(region_header_row, col).value)
            internal = SOURCE_REGION_TO_INTERNAL.get(header)
            if internal is not None:
                region_columns[internal] = col

    parsed: dict[str, dict[str, float]] = {}
    for row in range(start_row + 1, min(sheet.max_row, start_row + 220)):
        label = _cell_text(sheet.cell(row, 1).value)
        if not label:
            continue

        lowered = label.lower()
        if lowered.startswith("contents") and row > start_row + 6:
            break
        if lowered.startswith("total"):
            break

        canonical = PARTY_NAME_MAP.get(label.strip())
        if canonical is None:
            continue

        pct_value = sheet.cell(row + 1, 2).value
        if not isinstance(pct_value, (int, float)):
            continue

        region_values = {NATIONAL_KEY: _to_percentage(pct_value)}
        for region_name, col in region_columns.items():
            region_values[region_name] = _to_percentage_or_zero(sheet.cell(row + 1, col).value)
        parsed[canonical] = region_values

    for optional_party in ["Scottish National Party", "Plaid Cymru", "Other"]:
        parsed.setdefault(optional_party, {NATIONAL_KEY: 0.0})
        for region_name in SOURCE_REGION_TO_INTERNAL.values():
            parsed[optional_party].setdefault(region_name, 0.0)

    required = {
        "Conservative",
        "Labour",
        "Liberal Democrats",
        "Reform UK",
        "Green",
        "Scottish National Party",
        "Plaid Cymru",
        "Other",
    }
    missing = sorted(required - set(parsed.keys()))
    if missing:
        raise ValueError(f"Missing expected party rows in workbook: {missing}")

    return parsed


def parse_poll(workbook: Any, *, source_url: str, year_hint: int | None = None) -> ParsedPoll:
    year = _infer_year(source_url, fallback=year_hint)
    fieldwork_start, fieldwork_end, sample_size = _parse_cover_metadata(workbook, year)
    party_region_percentages = _parse_party_region_percentages(workbook)
    return ParsedPoll(
        sample_size=sample_size,
        fieldwork_start=fieldwork_start,
        fieldwork_end=fieldwork_end,
        party_region_percentages=party_region_percentages,
    )


def _find_existing_poll(db: Database, pollster_id: int, map_id: int, parsed: ParsedPoll) -> Poll | None:
    with db.session() as session:
        return session.execute(
            select(Poll).where(
                Poll.pollster_id == pollster_id,
                Poll.map_id == map_id,
                Poll.fieldwork_start == parsed.fieldwork_start,
                Poll.fieldwork_end == parsed.fieldwork_end,
                Poll.sample_size == parsed.sample_size,
            )
        ).scalar_one_or_none()


def build_import_plan(
    db: Database,
    *,
    xlsx_url: str = DEFAULT_XLSX_URL,
    map_name: str = DEFAULT_MAP_NAME,
    pollster_identifier: str = DEFAULT_POLLSTER_IDENTIFIER,
    year_hint: int | None = None,
) -> ImportPlan:
    poll_map = db.get_map_by_name(map_name)
    if poll_map is None:
        raise ValueError(f"Map not found: {map_name!r}")

    workbook = extract_workbook(xlsx_url)
    parsed = parse_poll(workbook, source_url=xlsx_url, year_hint=year_hint)

    pollster = db.get_pollster_by_identifier(pollster_identifier)
    pollster_exists = pollster is not None

    party_by_name = {party.name: party for party in db.get_all_parties()}
    missing_parties = [
        party_name
        for party_name in sorted(set(PARTY_NAME_MAP.values()))
        if party_name not in party_by_name
    ]
    if missing_parties:
        raise ValueError(
            "Missing parties in database (run party importer first): "
            f"{missing_parties}"
        )

    regions = db.get_regions_for_map(poll_map.id)
    region_ids_by_name: dict[str, int] = {region.name: region.id for region in regions}

    rows: list[PlannedPollRow] = []
    for party_name, region_values in parsed.party_region_percentages.items():
        rows.append(
            PlannedPollRow(
                party_id=party_by_name[party_name].id,
                party_name=party_name,
                region_id=None,
                region_name="National",
                percentage=region_values.get(NATIONAL_KEY, 0.0),
            )
        )

        for region in regions:
            rows.append(
                PlannedPollRow(
                    party_id=party_by_name[party_name].id,
                    party_name=party_name,
                    region_id=region.id,
                    region_name=region.name,
                    percentage=region_values.get(region.name, 0.0),
                )
            )

    existing_poll = (
        _find_existing_poll(db, pollster.id, poll_map.id, parsed)
        if pollster is not None
        else None
    )

    return ImportPlan(
        pollster_identifier=pollster_identifier,
        pollster_name=(pollster.name if pollster else "Survation"),
        pollster_id=(pollster.id if pollster else None),
        pollster_exists=pollster_exists,
        regions_mapping="",
        map_id=poll_map.id,
        map_name=poll_map.name,
        source_url=xlsx_url,
        parsed=parsed,
        poll_id=(existing_poll.id if existing_poll else None),
        poll_exists=existing_poll is not None,
        rows=rows,
    )


def commit_import_plan(
    db: Database,
    plan: ImportPlan,
    *,
    replace_rows: bool = False,
) -> PollImportResult:
    if plan.pollster_exists:
        pollster = db.get_pollster_by_identifier(plan.pollster_identifier)
        if pollster is None:
            raise ValueError("Pollster lookup failed during commit")
        pollster_id = pollster.id
        created_pollster = False
    else:
        created = db.add_pollster(
            name=plan.pollster_name,
            identifier=plan.pollster_identifier,
            weight=1.0,
            regions_mapping=plan.regions_mapping,
        )
        pollster_id = created.id
        created_pollster = True

    existing_poll = _find_existing_poll(db, pollster_id, plan.map_id, plan.parsed)
    if existing_poll is None:
        created_poll = db.add_poll(
            pollster_id=pollster_id,
            map_id=plan.map_id,
            fieldwork_start=plan.parsed.fieldwork_start,
            fieldwork_end=plan.parsed.fieldwork_end,
            sample_size=plan.parsed.sample_size,
            source_url=plan.source_url,
        )
        poll_id = created_poll.id
        created_poll_row = True
    else:
        with db.session() as session:
            db_poll = session.get(Poll, existing_poll.id)
            if db_poll is not None and db_poll.source_url != plan.source_url:
                db_poll.source_url = plan.source_url
        poll_id = existing_poll.id
        created_poll_row = False

    with db.session() as session:
        existing_rows = session.execute(
            select(PollRow).where(PollRow.poll_id == poll_id)
        ).scalars().all()

        if existing_rows and not replace_rows:
            return PollImportResult(
                created_pollster=created_pollster,
                created_poll=created_poll_row,
                poll_id=poll_id,
                inserted_rows=0,
                replaced_rows=0,
                skipped_existing_rows=True,
            )

        replaced_rows = 0
        if existing_rows and replace_rows:
            for existing_row in existing_rows:
                session.delete(existing_row)
                replaced_rows += 1

        for row in plan.rows:
            session.add(
                PollRow(
                    poll_id=poll_id,
                    region_id=row.region_id,
                    party_id=row.party_id,
                    percentage=row.percentage,
                )
            )

    return PollImportResult(
        created_pollster=created_pollster,
        created_poll=created_poll_row,
        poll_id=poll_id,
        inserted_rows=len(plan.rows),
        replaced_rows=replaced_rows,
        skipped_existing_rows=False,
    )


def _cli_preview(plan: ImportPlan) -> None:
    print(
        "Parsed poll: "
        f"fieldwork={plan.parsed.fieldwork_start} to {plan.parsed.fieldwork_end}, "
        f"sample={plan.parsed.sample_size}"
    )
    if plan.pollster_exists:
        print(f"pollster exists: {plan.pollster_identifier}")
    else:
        print(f"[dry-run] would create pollster: {plan.pollster_identifier}")

    if plan.poll_exists and plan.poll_id is not None:
        print(f"poll exists: {plan.poll_id}")
    else:
        print("[dry-run] would create poll")

    for row in plan.rows:
        print(
            "[dry-run] would insert row: "
            f"party={row.party_name}, region={row.region_name}, "
            f"region_id={row.region_id}, pct={row.percentage:.2f}"
        )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx-url", default=DEFAULT_XLSX_URL)
    parser.add_argument("--map-name", default=DEFAULT_MAP_NAME)
    parser.add_argument("--pollster-identifier", default=DEFAULT_POLLSTER_IDENTIFIER)
    parser.add_argument("--year-hint", type=int, default=None)
    parser.add_argument(
        "--replace-rows",
        action="store_true",
        help="Delete existing rows for the poll before inserting new ones",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    db = Database()
    print(f"Fetching XLSX: {args.xlsx_url}")

    plan = build_import_plan(
        db,
        xlsx_url=args.xlsx_url,
        map_name=args.map_name,
        pollster_identifier=args.pollster_identifier,
        year_hint=args.year_hint,
    )

    if args.dry_run:
        _cli_preview(plan)
        return

    result = commit_import_plan(db, plan, replace_rows=args.replace_rows)
    if result.created_pollster:
        print(f"created pollster: {args.pollster_identifier}")
    else:
        print(f"pollster exists: {args.pollster_identifier}")

    if result.created_poll:
        print(f"created poll: {result.poll_id}")
    else:
        print(f"poll exists: {result.poll_id}")

    if result.skipped_existing_rows:
        print(
            f"poll {result.poll_id} already has rows; "
            "use --replace-rows to overwrite"
        )
    else:
        if result.replaced_rows:
            print(f"deleted existing rows: {result.replaced_rows}")
        print(f"inserted poll rows: {result.inserted_rows}")


if __name__ == "__main__":
    main()
