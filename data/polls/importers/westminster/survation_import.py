#!/usr/bin/env python3
"""Import a Survation poll directly from an XLSX URL."""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import select

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent.parent))

from db import Database
from models import Poll, PollRow, Pollster
from polls.importers.types import PollImportResult

DEFAULT_XLSX_URL = "https://cdn.survation.com/wp-content/uploads/2026/01/30163059/Survation_Voter_Intention_Jan_2026.xlsx"
DEFAULT_MAP_NAME = "UK Constituencies post 2022"
DEFAULT_POLLSTER_IDENTIFIER = "survation"
NATIONAL_KEY = "__national__"

SOURCE_REGION_TO_INTERNAL = {
    "East Midlands": "East Midlands",
    "East of England": "East of England",
    "London": "London",
    "North East": "North East England",
    "North West": "North West England",
    "South East": "South East England",
    "South West": "South West England",
    "West Midlands": "West Midlands",
    "Yorkshire and The Humber": "Yorkshire and The Humber",
    "Scotland": "Scotland",
    "Wales": "Wales",
    "Northern Ireland": "Northern Ireland",
}

PARTY_NAME_MAP = {
    "Conservative": "Conservative",
    "Labour": "Labour",
    "Liberal Democrats": "Liberal Democrats",
    "Liberal Democrat": "Liberal Democrats",
    "Reform UK": "Reform UK",
    "Green": "Green",
    "Green Party": "Green",
    "Scottish National Party": "Scottish National Party",
    "Scottish National Party (SNP)": "Scottish National Party",
    "SNP": "Scottish National Party",
    "Plaid Cymru": "Plaid Cymru",
    "Another party": "Other",
    "Other": "Other",
}


class ParsedPoll(BaseModel):
    """Parsed poll data extracted from source."""

    sample_size: int = Field(gt=0)
    fieldwork_start: date
    fieldwork_end: date
    party_region_percentages: dict[str, dict[str, float]]


class PlannedPollRow(BaseModel):
    """A single poll row planned for DB insertion."""

    party_id: int
    party_name: str
    region_id: int | None
    region_name: str
    percentage: float = Field(ge=0, le=100)


class ImportPlan(BaseModel):
    """Full import plan: pollster, poll metadata, and rows."""

    pollster_identifier: str
    pollster_name: str
    pollster_id: int | None
    pollster_exists: bool
    regions_mapping: str
    map_id: int
    map_name: str
    source_url: str
    parsed: ParsedPoll
    poll_id: int | None
    poll_exists: bool
    rows: list[PlannedPollRow]


def _month_number(month_text: str) -> int | None:
    """Convert a month name string to its integer number (1–12).

    Args:
        month_text: Month name, either full (e.g. "January") or abbreviated
            (e.g. "Jan"). Case-insensitive; trailing periods are stripped.

    Returns:
        Integer month number (1–12), or None if the string is not recognised.
    """
    month_map = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }
    return month_map.get(month_text.strip().lower().rstrip("."))


def _infer_year(url: str, fallback: int | None = None) -> int:
    """Extract a four-digit year (20xx) from a URL string.

    First looks for a year component appearing as a path segment (e.g.
    ``/2026/``); if that is absent, falls back to the first bare ``20xx``
    occurrence in the URL.

    Args:
        url: URL string to search, typically the XLSX download URL.
        fallback: Year to return when no year can be found in the URL.
            Optional; if omitted and no year is found, a ValueError is raised.

    Returns:
        Four-digit year as an integer.

    Raises:
        ValueError: If no year is found in the URL and no fallback is provided.
    """
    match = re.search(r"/(20\d{2})/", url)
    if match:
        return int(match.group(1))
    matches = re.findall(r"(20\d{2})", url)
    if matches:
        return int(matches[0])
    if fallback is not None:
        return fallback
    raise ValueError("Could not infer year from URL")


def _parse_fieldwork(fieldwork_text: str, default_year: int | None = None) -> tuple[date, date]:
    """Parse a fieldwork date-range string into a (start, end) date pair.

    Handles several formats found in Survation cover sheets, normalising
    Unicode dashes and ordinal suffixes before matching:

    - Same-month with explicit year: ``"3-5 January 2026"``
    - Same-month without year: ``"3-5 January"`` (requires ``default_year``)
    - Cross-month without year: ``"30 Jan - 1 Feb"`` (requires ``default_year``)
    - Cross-month with explicit year: ``"30 Jan - 1 Feb 2026"``

    When a cross-month range spans a year boundary (e.g. ``"31 Dec - 2 Jan"``),
    the start year is decremented by one automatically.

    Args:
        fieldwork_text: Raw fieldwork date string as it appears in the workbook.
        default_year: Year to use when the string contains no explicit year.
            Optional; if omitted and no year is present in the string, parsing
            will fall through and raise a ValueError.

    Returns:
        A tuple of ``(fieldwork_start, fieldwork_end)`` as :class:`datetime.date`
        objects.

    Raises:
        ValueError: If the month names cannot be parsed or no pattern matches
            the input string.
    """
    normalized = re.sub(r"\s+", " ", fieldwork_text.strip().replace("–", "-").replace("—", "-"))
    normalized = re.sub(r"(\d)(st|nd|rd|th)", r"\1", normalized, flags=re.IGNORECASE)

    pattern_same_month = re.compile(r"(\d{1,2})\s*-\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})")
    match = pattern_same_month.search(normalized)
    if match:
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        year = int(match.group(4))
        if month is None:
            raise ValueError(f"Could not parse month in fieldwork: {fieldwork_text!r}")
        return date(year, month, day_start), date(year, month, day_end)

    pattern_no_year = re.compile(r"(\d{1,2})\s*-\s*(\d{1,2})\s+([A-Za-z]+)")
    match = pattern_no_year.search(normalized)
    if match and default_year is not None:
        day_start = int(match.group(1))
        day_end = int(match.group(2))
        month = _month_number(match.group(3))
        if month is None:
            raise ValueError(f"Could not parse month in fieldwork: {fieldwork_text!r}")
        return date(default_year, month, day_start), date(default_year, month, day_end)

    pattern_cross_no_year = re.compile(r"(\d{1,2})\s+([A-Za-z]+)\s*-\s*(\d{1,2})\s+([A-Za-z]+)")
    match = pattern_cross_no_year.search(normalized)
    if match and default_year is not None:
        day_start = int(match.group(1))
        month_start = _month_number(match.group(2))
        day_end = int(match.group(3))
        month_end = _month_number(match.group(4))
        if month_start is None or month_end is None:
            raise ValueError(f"Could not parse months in fieldwork: {fieldwork_text!r}")
        year_start = default_year - 1 if month_start > month_end else default_year
        return date(year_start, month_start, day_start), date(default_year, month_end, day_end)

    pattern_cross_month = re.compile(
        r"(\d{1,2})\s+([A-Za-z]+)\s*-\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})"
    )
    match = pattern_cross_month.search(normalized)
    if match:
        day_start = int(match.group(1))
        month_start = _month_number(match.group(2))
        day_end = int(match.group(3))
        month_end = _month_number(match.group(4))
        year_end = int(match.group(5))
        if month_start is None or month_end is None:
            raise ValueError(f"Could not parse months in fieldwork: {fieldwork_text!r}")
        year_start = year_end - 1 if month_start > month_end else year_end
        return date(year_start, month_start, day_start), date(year_end, month_end, day_end)

    raise ValueError(f"Could not parse fieldwork string: {fieldwork_text!r}")


def extract_workbook(xlsx_url: str) -> Any:
    """Download an XLSX file from a URL and return an openpyxl workbook.

    Sends a browser-like User-Agent header to avoid being blocked by CDN
    servers. Validates that the response payload begins with the ZIP magic
    bytes (``PK``) expected of a valid XLSX file.

    Args:
        xlsx_url: Fully-qualified URL of the XLSX file to download.

    Returns:
        An openpyxl ``Workbook`` object loaded with ``data_only=True`` so
        that formula results rather than formula strings are read.

    Raises:
        ValueError: If the downloaded payload does not appear to be a valid
            XLSX file (missing ZIP magic bytes).
        urllib.error.URLError: If the network request fails.
    """
    req = Request(xlsx_url, headers={"User-Agent": "Mozilla/5.0 (compatible; poll-importer/1.0)"})
    with urlopen(req, timeout=50) as response:
        payload = response.read()

    if not payload.startswith(b"PK"):
        raise ValueError(f"Could not fetch XLSX payload from URL: {xlsx_url}")

    return load_workbook(filename=BytesIO(payload), data_only=True)


def _cell_text(value: object) -> str:
    """Return the stripped string representation of a cell value.

    Args:
        value: Raw cell value from openpyxl (may be None, int, float, or str).

    Returns:
        Stripped string, or an empty string if the value is None.
    """
    if value is None:
        return ""
    return str(value).strip()


def _to_percentage(value: object) -> float:
    """Convert a cell value to a percentage float in the range 0–100.

    Values already in the 0–100 range are returned as-is; values in the
    0–1 range (i.e. stored as decimals) are multiplied by 100. The result
    is rounded to the nearest whole number.

    Args:
        value: Raw cell value, expected to be numeric (int or float).

    Returns:
        Percentage as a float (e.g. ``42.0``).

    Raises:
        ValueError: If the value is None.
    """
    if value is None:
        raise ValueError("Encountered empty percentage cell")
    number = float(value)  # type: ignore[arg-type]
    pct = number * 100.0 if 0.0 <= number <= 1.0 else number
    return float(int(round(pct)))


def _to_percentage_or_zero(value: object) -> float:
    """Convert a cell value to a percentage float, returning 0.0 for blank cells.

    Treats None and dash-like strings (``"-"``, ``"–"``, ``"—"``, etc.) as
    zero rather than raising an error. Delegates to :func:`_to_percentage` for
    all other values.

    Args:
        value: Raw cell value from openpyxl. May be None, a numeric type, or a
            string placeholder such as ``"-"`` indicating a missing value.

    Returns:
        Percentage as a float, or ``0.0`` if the cell is blank or a dash.
    """
    if value is None:
        return 0.0
    if isinstance(value, str):
        cleaned = value.strip()
        if cleaned in {"", "-", "–", "—", "- ", "-\u00a0"}:
            return 0.0
    return _to_percentage(value)


def _cover_sheet(workbook: Any) -> Any:
    """Return the cover/methodology sheet from a Survation workbook.

    Looks for a sheet whose name contains both "cover" and "method"
    (case-insensitive). Falls back to the first sheet in the workbook if no
    such sheet is found.

    Args:
        workbook: An openpyxl ``Workbook`` object.

    Returns:
        The matching openpyxl ``Worksheet``.
    """
    for name in workbook.sheetnames:
        if "cover" in name.lower() and "method" in name.lower():
            return workbook[name]
    return workbook[workbook.sheetnames[0]]


def _parse_cover_metadata(workbook: Any, default_year: int) -> tuple[date, date, int]:
    """Parse fieldwork dates and sample size from the cover sheet.

    Scans the first 80 rows of column A looking for the labels
    ``"Fieldwork Dates"`` and ``"Sample Size"``; the values are expected in
    the row immediately below each label.

    Args:
        workbook: An openpyxl ``Workbook`` object for the Survation XLSX.
        default_year: Year to pass to :func:`_parse_fieldwork` when the
            fieldwork date string contains no explicit year.

    Returns:
        A tuple of ``(fieldwork_start, fieldwork_end, sample_size)`` where
        the dates are :class:`datetime.date` objects and ``sample_size`` is a
        positive integer.

    Raises:
        ValueError: If fieldwork dates or sample size cannot be located in the
            cover sheet.
    """
    cover = _cover_sheet(workbook)

    fieldwork_raw = None
    sample_size = None

    for row in range(1, 80):
        label = _cell_text(cover.cell(row, 1).value).lower()
        if "fieldwork dates" in label:
            fieldwork_raw = _cell_text(cover.cell(row + 1, 1).value)
        if label == "sample size":
            candidate = cover.cell(row + 1, 1).value
            if isinstance(candidate, (int, float)):
                sample_size = int(round(float(candidate)))
            else:
                digits = re.sub(r"[^0-9]", "", _cell_text(candidate))
                if digits:
                    sample_size = int(digits)

    if not fieldwork_raw:
        raise ValueError("Fieldwork dates not found in cover sheet")
    if sample_size is None:
        raise ValueError("Sample size not found in cover sheet")

    fieldwork_start, fieldwork_end = _parse_fieldwork(fieldwork_raw, default_year=default_year)
    return fieldwork_start, fieldwork_end, sample_size


def _tables_sheet(workbook: Any) -> Any:
    """Locate and return the data tables sheet from a Survation workbook.

    Applies a priority search across sheet names:

    1. Exact match for ``"tables"`` or ``"table"`` (case-insensitive).
    2. Sheet name that starts with ``"tables"``.
    3. Sheet name that contains ``"table"`` but not ``"contents"``.

    Args:
        workbook: An openpyxl ``Workbook`` object.

    Returns:
        The matching openpyxl ``Worksheet``.

    Raises:
        ValueError: If no suitable tables sheet can be found.
    """
    for name in workbook.sheetnames:
        lowered = name.strip().lower()
        if lowered in {"tables", "table"}:
            return workbook[name]
    for name in workbook.sheetnames:
        lowered = name.strip().lower()
        if lowered.startswith("tables"):
            return workbook[name]
    for name in workbook.sheetnames:
        lowered = name.strip().lower()
        if "table" in lowered and "contents" not in lowered:
            return workbook[name]
    raise ValueError("Could not find tables sheet")


def _find_vi_table_start(sheet: Any) -> int:
    """Scan a tables sheet for the voting-intention table header row.

    Searches column A for rows that contain a ``table_`` marker together with
    references to a general/Westminster election and a voting question. Among
    matching rows, prefers those whose immediately following row mentions
    "undecided" and "remove" (indicating the de-undecided base). Returns the
    last preferred match, or the last plain match if none are preferred.

    Args:
        sheet: An openpyxl ``Worksheet`` from a Survation XLSX tables sheet.

    Returns:
        Row number (1-based) of the voting-intention table header.

    Raises:
        ValueError: If no matching table header can be located in the sheet.
    """
    starts: list[int] = []
    preferred: list[int] = []

    for row in range(1, sheet.max_row + 1):
        header = _cell_text(sheet.cell(row, 1).value)
        header_lower = header.lower()
        if "table_" not in header_lower:
            continue
        if "general election" not in header_lower and "westminster election" not in header_lower:
            continue
        if "for which party" not in header_lower or "vote" not in header_lower:
            continue

        starts.append(row)
        base = _cell_text(sheet.cell(row + 1, 1).value).lower()
        if "undecided" in base and "remove" in base:
            preferred.append(row)

    if preferred:
        return preferred[-1]
    if starts:
        return starts[-1]
    raise ValueError("Could not locate voting intention table in Survation workbook")


def _parse_party_region_percentages(workbook: Any) -> dict[str, dict[str, float]]:
    """Extract party vote-share percentages by region from the tables sheet.

    Locates the voting-intention table, identifies the region column headers,
    then reads the percentage value from the row below each party label.
    Party names are normalised via ``PARTY_NAME_MAP``; region names are
    normalised via ``SOURCE_REGION_TO_INTERNAL``.

    Optional parties (SNP, Plaid Cymru, Other) default to 0.0 if absent.
    All region keys for those parties are also defaulted to 0.0.

    Args:
        workbook: An openpyxl ``Workbook`` object for the Survation XLSX.

    Returns:
        A nested dict mapping canonical party name → region name →
        percentage float (0–100). The national figure is stored under the
        ``NATIONAL_KEY`` (``"__national__"``) key.

    Raises:
        ValueError: If any of the required party rows are missing from the
            parsed data.
    """
    sheet = _tables_sheet(workbook)
    start_row = _find_vi_table_start(sheet)

    region_header_row = None
    for row in range(start_row + 1, min(sheet.max_row, start_row + 12)):
        headers = {_cell_text(sheet.cell(row, col).value) for col in range(11, 40)}
        if any(header in SOURCE_REGION_TO_INTERNAL for header in headers):
            region_header_row = row
            break

    region_columns: dict[str, int] = {}
    if region_header_row is not None:
        for col in range(11, 45):
            header = _cell_text(sheet.cell(region_header_row, col).value)
            internal = SOURCE_REGION_TO_INTERNAL.get(header)
            if internal is not None:
                region_columns[internal] = col

    parsed: dict[str, dict[str, float]] = {}
    for row in range(start_row + 1, min(sheet.max_row, start_row + 220)):
        label = _cell_text(sheet.cell(row, 1).value)
        if not label:
            continue

        lowered = label.lower()
        if lowered.startswith("contents") and row > start_row + 6:
            break
        if lowered.startswith("total"):
            break

        canonical = PARTY_NAME_MAP.get(label.strip())
        if canonical is None:
            continue

        pct_value = sheet.cell(row + 1, 2).value
        if not isinstance(pct_value, (int, float)):
            continue

        region_values = {NATIONAL_KEY: _to_percentage(pct_value)}
        for region_name, col in region_columns.items():
            region_values[region_name] = _to_percentage_or_zero(sheet.cell(row + 1, col).value)
        parsed[canonical] = region_values

    for optional_party in ["Scottish National Party", "Plaid Cymru", "Other"]:
        parsed.setdefault(optional_party, {NATIONAL_KEY: 0.0})
        for region_name in SOURCE_REGION_TO_INTERNAL.values():
            parsed[optional_party].setdefault(region_name, 0.0)

    required = {
        "Conservative",
        "Labour",
        "Liberal Democrats",
        "Reform UK",
        "Green",
        "Scottish National Party",
        "Plaid Cymru",
        "Other",
    }
    missing = sorted(required - set(parsed.keys()))
    if missing:
        raise ValueError(f"Missing expected party rows in workbook: {missing}")

    return parsed


def parse_poll(workbook: Any, *, source_url: str, year_hint: int | None = None) -> ParsedPoll:
    """Parse a downloaded Survation workbook into a :class:`ParsedPoll`.

    Combines cover-sheet metadata (fieldwork dates, sample size) with the
    party-by-region percentages extracted from the tables sheet.

    Args:
        workbook: An openpyxl ``Workbook`` object for the Survation XLSX.
        source_url: The URL from which the workbook was downloaded. Used to
            infer the poll year when the fieldwork string omits it.
        year_hint: Explicit fallback year to use if the year cannot be inferred
            from ``source_url``. Optional.

    Returns:
        A :class:`ParsedPoll` containing sample size, fieldwork dates, and the
        full party-region percentage mapping.

    Raises:
        ValueError: If year, fieldwork dates, sample size, or required party
            rows cannot be determined from the workbook.
    """
    year = _infer_year(source_url, fallback=year_hint)
    fieldwork_start, fieldwork_end, sample_size = _parse_cover_metadata(workbook, year)
    party_region_percentages = _parse_party_region_percentages(workbook)
    return ParsedPoll(
        sample_size=sample_size,
        fieldwork_start=fieldwork_start,
        fieldwork_end=fieldwork_end,
        party_region_percentages=party_region_percentages,
    )


def _find_existing_poll(db: Database, pollster_id: int, map_id: int, parsed: ParsedPoll) -> Poll | None:
    """Query the database for an existing poll matching the parsed metadata.

    Matches on pollster, map, fieldwork start/end dates, and sample size.

    Args:
        db: Active :class:`Database` instance.
        pollster_id: Primary key of the pollster row.
        map_id: Primary key of the electoral map row.
        parsed: Parsed poll data containing the fieldwork dates and sample size
            to match against.

    Returns:
        The matching :class:`Poll` ORM object, or None if no match is found.
    """
    with db.session() as session:
        return session.execute(
            select(Poll).where(
                Poll.pollster_id == pollster_id,
                Poll.map_id == map_id,
                Poll.fieldwork_start == parsed.fieldwork_start,
                Poll.fieldwork_end == parsed.fieldwork_end,
                Poll.sample_size == parsed.sample_size,
            )
        ).scalar_one_or_none()


def build_import_plan(
    db: Database,
    *,
    xlsx_url: str = DEFAULT_XLSX_URL,
    map_name: str = DEFAULT_MAP_NAME,
    pollster_identifier: str = DEFAULT_POLLSTER_IDENTIFIER,
    year_hint: int | None = None,
) -> ImportPlan:
    """Build an :class:`ImportPlan` describing what would be written to the DB.

    Downloads and parses the XLSX, resolves the map, pollster, and all parties
    against the current database state, and constructs the full list of
    :class:`PlannedPollRow` objects. Does **not** write anything to the
    database.

    Args:
        db: Active :class:`Database` instance.
        xlsx_url: URL of the Survation XLSX to import. Defaults to
            ``DEFAULT_XLSX_URL``.
        map_name: Name of the electoral map to associate the poll with.
            Defaults to ``DEFAULT_MAP_NAME``.
        pollster_identifier: Slug used to look up or create the pollster row.
            Defaults to ``DEFAULT_POLLSTER_IDENTIFIER``.
        year_hint: Explicit fallback year for fieldwork date parsing when the
            year cannot be inferred from the URL. Optional.

    Returns:
        An :class:`ImportPlan` containing pollster/poll existence flags and
        the full list of rows to insert.

    Raises:
        ValueError: If the named map is not found in the database, or if any
            party in the parsed poll data is missing from the database.
    """
    poll_map = db.get_map_by_name(map_name)
    if poll_map is None:
        raise ValueError(f"Map not found: {map_name!r}")

    workbook = extract_workbook(xlsx_url)
    parsed = parse_poll(workbook, source_url=xlsx_url, year_hint=year_hint)

    pollster = db.get_pollster_by_identifier(pollster_identifier)
    pollster_exists = pollster is not None

    party_by_name = {party.name: party for party in db.get_all_parties()}
    missing_parties = [
        party_name
        for party_name in sorted(set(PARTY_NAME_MAP.values()))
        if party_name not in party_by_name
    ]
    if missing_parties:
        raise ValueError(
            "Missing parties in database (run party importer first): "
            f"{missing_parties}"
        )

    regions = db.get_regions_for_map(poll_map.id)
    region_ids_by_name: dict[str, int] = {region.name: region.id for region in regions}

    rows: list[PlannedPollRow] = []
    for party_name, region_values in parsed.party_region_percentages.items():
        rows.append(
            PlannedPollRow(
                party_id=party_by_name[party_name].id,
                party_name=party_name,
                region_id=None,
                region_name="National",
                percentage=region_values.get(NATIONAL_KEY, 0.0),
            )
        )

        for region in regions:
            rows.append(
                PlannedPollRow(
                    party_id=party_by_name[party_name].id,
                    party_name=party_name,
                    region_id=region.id,
                    region_name=region.name,
                    percentage=region_values.get(region.name, 0.0),
                )
            )

    existing_poll = (
        _find_existing_poll(db, pollster.id, poll_map.id, parsed)
        if pollster is not None
        else None
    )

    return ImportPlan(
        pollster_identifier=pollster_identifier,
        pollster_name=(pollster.name if pollster else "Survation"),
        pollster_id=(pollster.id if pollster else None),
        pollster_exists=pollster_exists,
        regions_mapping="",
        map_id=poll_map.id,
        map_name=poll_map.name,
        source_url=xlsx_url,
        parsed=parsed,
        poll_id=(existing_poll.id if existing_poll else None),
        poll_exists=existing_poll is not None,
        rows=rows,
    )


def commit_import_plan(
    db: Database,
    plan: ImportPlan,
    *,
    replace_rows: bool = False,
) -> PollImportResult:
    """Execute an :class:`ImportPlan`, writing records to the database.

    Creates the pollster and poll rows if they do not already exist. If the
    poll already has rows and ``replace_rows`` is False, the function returns
    immediately without inserting new rows. If ``replace_rows`` is True,
    existing rows are deleted before the new rows are inserted.

    Args:
        db: Active :class:`Database` instance.
        plan: The :class:`ImportPlan` produced by :func:`build_import_plan`.
        replace_rows: When True, delete any existing :class:`PollRow` records
            for the poll before inserting the new rows. Defaults to False.

    Returns:
        A :class:`PollImportResult` summarising what was created, inserted, or
        skipped.

    Raises:
        ValueError: If the pollster lookup fails during commit despite the plan
            indicating the pollster exists.
    """
    if plan.pollster_exists:
        pollster = db.get_pollster_by_identifier(plan.pollster_identifier)
        if pollster is None:
            raise ValueError("Pollster lookup failed during commit")
        pollster_id = pollster.id
        created_pollster = False
    else:
        created = db.add_pollster(
            name=plan.pollster_name,
            identifier=plan.pollster_identifier,
            weight=1.0,
            regions_mapping=plan.regions_mapping,
        )
        pollster_id = created.id
        created_pollster = True

    existing_poll = _find_existing_poll(db, pollster_id, plan.map_id, plan.parsed)
    if existing_poll is None:
        created_poll = db.add_poll(
            pollster_id=pollster_id,
            map_id=plan.map_id,
            fieldwork_start=plan.parsed.fieldwork_start,
            fieldwork_end=plan.parsed.fieldwork_end,
            sample_size=plan.parsed.sample_size,
            source_url=plan.source_url,
        )
        poll_id = created_poll.id
        created_poll_row = True
    else:
        with db.session() as session:
            db_poll = session.get(Poll, existing_poll.id)
            if db_poll is not None and db_poll.source_url != plan.source_url:
                db_poll.source_url = plan.source_url
        poll_id = existing_poll.id
        created_poll_row = False

    with db.session() as session:
        existing_rows = session.execute(
            select(PollRow).where(PollRow.poll_id == poll_id)
        ).scalars().all()

        if existing_rows and not replace_rows:
            return PollImportResult(
                created_pollster=created_pollster,
                created_poll=created_poll_row,
                poll_id=poll_id,
                inserted_rows=0,
                replaced_rows=0,
                skipped_existing_rows=True,
            )

        replaced_rows = 0
        if existing_rows and replace_rows:
            for existing_row in existing_rows:
                session.delete(existing_row)
                replaced_rows += 1

        for row in plan.rows:
            session.add(
                PollRow(
                    poll_id=poll_id,
                    region_id=row.region_id,
                    party_id=row.party_id,
                    percentage=row.percentage,
                )
            )

    return PollImportResult(
        created_pollster=created_pollster,
        created_poll=created_poll_row,
        poll_id=poll_id,
        inserted_rows=len(plan.rows),
        replaced_rows=replaced_rows,
        skipped_existing_rows=False,
    )


def _cli_preview(plan: ImportPlan) -> None:
    """Print a dry-run preview of an import plan to stdout.

    Displays parsed fieldwork dates and sample size, pollster existence status,
    poll existence status, and one line per planned poll row showing party,
    region, region ID, and percentage.

    Args:
        plan: The :class:`ImportPlan` to preview.
    """
    print(
        "Parsed poll: "
        f"fieldwork={plan.parsed.fieldwork_start} to {plan.parsed.fieldwork_end}, "
        f"sample={plan.parsed.sample_size}"
    )
    if plan.pollster_exists:
        print(f"pollster exists: {plan.pollster_identifier}")
    else:
        print(f"[dry-run] would create pollster: {plan.pollster_identifier}")

    if plan.poll_exists and plan.poll_id is not None:
        print(f"poll exists: {plan.poll_id}")
    else:
        print("[dry-run] would create poll")

    for row in plan.rows:
        print(
            "[dry-run] would insert row: "
            f"party={row.party_name}, region={row.region_name}, "
            f"region_id={row.region_id}, pct={row.percentage:.2f}"
        )


def main() -> None:
    """CLI entry point for the Survation poll importer.

    Parses command-line arguments, builds an import plan (downloading and
    parsing the XLSX), and either prints a dry-run preview or commits the plan
    to the database.

    Command-line arguments:
        --xlsx-url: URL of the Survation XLSX file to import.
            Defaults to ``DEFAULT_XLSX_URL``.
        --map-name: Name of the electoral map to associate the poll with.
            Defaults to ``DEFAULT_MAP_NAME``.
        --pollster-identifier: Slug for the pollster row lookup/creation.
            Defaults to ``DEFAULT_POLLSTER_IDENTIFIER``.
        --year-hint: Optional integer year used as a fallback when the poll
            year cannot be inferred from the XLSX URL.
        --replace-rows: Flag; if set, existing poll rows are deleted before
            inserting new ones.
        --dry-run: Flag; if set, prints a preview of what would be imported
            without writing to the database.
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx-url", default=DEFAULT_XLSX_URL)
    parser.add_argument("--map-name", default=DEFAULT_MAP_NAME)
    parser.add_argument("--pollster-identifier", default=DEFAULT_POLLSTER_IDENTIFIER)
    parser.add_argument("--year-hint", type=int, default=None)
    parser.add_argument(
        "--replace-rows",
        action="store_true",
        help="Delete existing rows for the poll before inserting new ones",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    db = Database()
    print(f"Fetching XLSX: {args.xlsx_url}")

    plan = build_import_plan(
        db,
        xlsx_url=args.xlsx_url,
        map_name=args.map_name,
        pollster_identifier=args.pollster_identifier,
        year_hint=args.year_hint,
    )

    if args.dry_run:
        _cli_preview(plan)
        return

    result = commit_import_plan(db, plan, replace_rows=args.replace_rows)
    if result.created_pollster:
        print(f"created pollster: {args.pollster_identifier}")
    else:
        print(f"pollster exists: {args.pollster_identifier}")

    if result.created_poll:
        print(f"created poll: {result.poll_id}")
    else:
        print(f"poll exists: {result.poll_id}")

    if result.skipped_existing_rows:
        print(
            f"poll {result.poll_id} already has rows; "
            "use --replace-rows to overwrite"
        )
    else:
        if result.replaced_rows:
            print(f"deleted existing rows: {result.replaced_rows}")
        print(f"inserted poll rows: {result.inserted_rows}")


if __name__ == "__main__":
    main()
